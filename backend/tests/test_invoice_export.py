"""Tests for `app.invoice_export` — TZ §8.1–§8.3 & §9.

Exercises openpyxl row-shift / merge handling against the real
`nakladna_template.xlsx` template — no DB needed.
"""
import os

from openpyxl import load_workbook

from app.invoice_export import (
    ROW_FIN_POST, ROW_FIN_NAME, TEMPLATE_SLOTS, TEMPLATE_LAST_ROW,
    adjust_item_rows,
)

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__), "..", "app", "nakladna_template.xlsx"
)


def _load():
    return load_workbook(TEMPLATE_PATH)


# ── Template integrity ────────────────────────────────────────────────────

def test_template_baseline():
    """Template starts with «Всього» at A25, fin merge A41:C43, print A1:J44."""
    wb = _load()
    ws = wb.active
    assert ws["A25"].value == "Всього"
    assert any(str(mr) == "A41:C43" for mr in ws.merged_cells.ranges), \
        "fin-post merge A41:C43 missing"
    assert ws.print_area == "'Накладна440_25'!$A$1:$J$44"


def test_no_defined_names_leak():
    """TZ regression: VBA-linked defined names cause «We found a problem» in
    Excel. Template must be clean.
    """
    wb = _load()
    assert list(wb.defined_names.keys()) == []


# ── adjust_item_rows ──────────────────────────────────────────────────────

def test_n_equals_5_no_shift():
    """N=5 baseline: no rows inserted/deleted, layout unchanged."""
    wb = _load()
    ws = wb.active
    delta = adjust_item_rows(ws, 5)
    assert delta == 0
    assert ws["A25"].value == "Всього"
    assert ws.cell(ROW_FIN_POST, 1).value == "начальник фінансової служби"


def test_n_equals_12_inserts_7_rows():
    """N=12: 7 rows inserted; footer shifts down by 7."""
    wb = _load()
    ws = wb.active
    delta = adjust_item_rows(ws, 12)
    assert delta == 7
    assert ws["A32"].value == "Всього"
    fin_post_row = ROW_FIN_POST + delta   # 48
    fin_name_row = ROW_FIN_NAME + delta   # 50
    assert ws.cell(fin_post_row, 1).value == "начальник фінансової служби"
    merge_strs = [str(mr) for mr in ws.merged_cells.ranges]
    assert f"A{fin_post_row}:C{fin_name_row}" in merge_strs


def test_n_equals_1_deletes_4_rows():
    """N=1: 4 rows deleted; footer shifts up by 4."""
    wb = _load()
    ws = wb.active
    delta = adjust_item_rows(ws, 1)
    assert delta == -4
    assert ws["A21"].value == "Всього"
    fin_post_row = ROW_FIN_POST + delta   # 37
    fin_name_row = ROW_FIN_NAME + delta   # 39
    merge_strs = [str(mr) for mr in ws.merged_cells.ranges]
    assert f"A{fin_post_row}:C{fin_name_row}" in merge_strs


def test_inserted_rows_inherit_item_row_style():
    """Newly-inserted rows must carry forward the item-row borders (col A has
    no left border in original, col C has no bottom — both must persist).
    """
    wb = _load()
    ws = wb.active
    orig_a20_left = ws.cell(20, 1).border.left.style
    orig_c20_bottom = ws.cell(20, 3).border.bottom.style

    adjust_item_rows(ws, 12)
    # First newly-inserted row at row 25
    assert ws.cell(25, 1).border.left.style == orig_a20_left
    assert ws.cell(25, 3).border.bottom.style == orig_c20_bottom


def test_template_last_row_constant():
    """TEMPLATE_LAST_ROW must match the template's actual N=5 last row."""
    wb = _load()
    ws = wb.active
    assert TEMPLATE_LAST_ROW == 44
    assert ws.max_row >= TEMPLATE_LAST_ROW
    assert TEMPLATE_SLOTS == 5
