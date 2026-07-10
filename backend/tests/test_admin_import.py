"""Unit tests for admin import helpers — pure-logic only.

Verifies the column-map parsing for items and the op-type → (operation,
form) mapping for movements without touching a DB.

`app.routers.admin` imports the SQLAlchemy stack which requires DATABASE_URL.
We set bogus values before import so config validation passes; we never
construct a session in this test file.
"""
import os
os.environ.setdefault("DATABASE_URL", "sqlite:///:memory:")
os.environ.setdefault("SECRET_KEY", "test-only")

from openpyxl import Workbook  # noqa: E402

from app.routers.admin import (  # noqa: E402
    ITEMS_COLUMN_MAP, OP_TYPE_MAP, RECIPIENT_SKIP_VALUES, SERIAL_NONE_TOKENS,
    _find_items_header_row, _build_items_col_map, _parse_decimal,
    _normalize_serial, _build_person_lookup, _resolve_person,
)


def _xlsx_with_items_header():
    wb = Workbook()
    ws = wb.active
    # Headers in row 1
    headers = ["№", "Товар", "Код номер", "Серійний номер",
               "Од. виміру", "Вартість", "Кіл-сть", "Категорія",
               "Де знаходиться", "Видано", "Дата видачі"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="K-001")
    ws.cell(row=2, column=2, value="Ноутбук")
    return wb


def test_find_items_header_row():
    wb = _xlsx_with_items_header()
    ws = wb.active
    assert _find_items_header_row(ws) == 1


def test_build_items_col_map_maps_all_headers():
    wb = _xlsx_with_items_header()
    ws = wb.active
    col_map = _build_items_col_map(ws, 1)
    # All 9 headers map to fields
    assert set(col_map.values()) == set(ITEMS_COLUMN_MAP.values())


def test_find_items_header_row_returns_none_when_missing():
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="random")
    assert _find_items_header_row(ws) is None


def test_op_type_map_canonical_values():
    # Map produces the 2-axis values the new schema expects
    assert OP_TYPE_MAP["надходження"] == ("надходження", "акт")
    assert OP_TYPE_MAP["переміщення"] == ("переміщення", "накладна")
    assert OP_TYPE_MAP["внутрішнє переміщення"] == ("переміщення", "накладна")


def test_parse_decimal_handles_ua_locale():
    from decimal import Decimal
    assert _parse_decimal("1 234,56") == Decimal("1234.56")
    assert _parse_decimal("0") == Decimal("0")
    assert _parse_decimal(None) is None
    assert _parse_decimal("not a number") is None
    assert _parse_decimal(42) == Decimal("42")


def test_видано_column_in_items_map():
    # Header maps to a synthetic key that the importer resolves to a recipient FK
    assert ITEMS_COLUMN_MAP.get("Видано") == "issued_recipient"


def test_recipient_skip_values_include_warehouse():
    # «Склад» / «склад» / empty must be skipped (don't autocreate a Склад recipient)
    assert "склад" in RECIPIENT_SKIP_VALUES
    assert "" in RECIPIENT_SKIP_VALUES


def test_normalize_serial_placeholders_become_none():
    for token in ("б/н", "Б/Н", " б/н ", "-", "—", "н/д", "", None, "  "):
        assert _normalize_serial(token) is None, f"expected None for {token!r}"


def test_normalize_serial_preserves_real_values():
    assert _normalize_serial("SN-12345") == "SN-12345"
    assert _normalize_serial(" ABC123 ") == "ABC123"  # trimmed
    assert _normalize_serial("б/н X") == "б/н X"      # only exact placeholder normalizes
    assert _normalize_serial(42) == "42"


class _FakePerson:
    """Duck-typed stand-in for models.Person — models pulling SQLAlchemy
    require DATABASE_URL at import time, and we've already stubbed env
    higher up; this keeps the helper pure."""
    def __init__(self, pid, first_name="", last_name="", search_name=None):
        self.id = pid
        self.first_name = first_name
        self.last_name = last_name
        self.search_name = search_name


def test_person_lookup_matches_first_last_case_insensitive():
    p = _FakePerson(1, first_name="Petro", last_name="Ivanenko", search_name="Ivanenko P.")
    lookup = _build_person_lookup([p])

    assert _resolve_person("Petro Ivanenko", lookup) == 1
    assert _resolve_person("PETRO IVANENKO", lookup) == 1
    assert _resolve_person("petro ivanenko", lookup) == 1
    assert _resolve_person("Ivanenko Petro", lookup) == 1   # swapped order
    assert _resolve_person("Ivanenko P.", lookup) == 1       # existing search_name
    assert _resolve_person("Nobody", lookup) is None


def test_person_lookup_last_name_alone():
    p = _FakePerson(2, last_name="Petrenko")
    lookup = _build_person_lookup([p])
    assert _resolve_person("Petrenko", lookup) == 2
    assert _resolve_person("PETRENKO", lookup) == 2


def test_movements_response_shape_includes_orphans_key():
    # The endpoint contract: response dict has these keys so the UI can
    # surface orphan card numbers separately from unmatched persons.
    # Pure-logic assertion via the router module's constant/key layout —
    # no live DB. If a future refactor renames the key, this test breaks
    # visibly before the UI breaks silently.
    import inspect
    from app.routers import admin as admin_mod
    src = inspect.getsource(admin_mod.import_movements)
    assert '"orphan_card_nums"' in src
    assert '"unmatched_persons"' in src
    assert '"documents_created"' in src
    assert '"auto_balanced"' in src
    assert '"auto_balanced_count"' in src


def test_auto_balance_logic_present_for_internal_transfers():
    # The auto-balance branch: for op=="переміщення" rows with only one of
    # qty_in/qty_out set, we mirror to the other side so the receiving
    # unit's SUM(qty_in) matches the sending unit's SUM(qty_out).
    import inspect
    from app.routers import admin as admin_mod
    src = inspect.getsource(admin_mod.import_movements)
    assert 'op == "переміщення"' in src
    assert 'in→out' in src
    assert 'out→in' in src
