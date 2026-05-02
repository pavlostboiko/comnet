#!/usr/bin/env python3
"""
Import items from Excel file.

Usage:
    docker compose exec backend python import_items.py /app/items.xlsx
    docker compose exec backend python import_items.py /app/items.xlsx --dry-run
"""

import sys
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

sys.path.insert(0, '/app')
from app.database import SessionLocal
from app.models import Item

# Excel column name → Item field
COLUMN_MAP = {
    '№':               'number',
    'Товар':           'name',
    'Код номер':       'nomenclature_code',
    'Серійний номер':  'serial_number',
    'Од. виміру':      'unit_of_measure',
    'Вартість':        'price',
    'Кіл-сть':         'quantity',
    'Категорія':       'item_type',
    'Де знаходиться':  'notes',
}

SKIP_COLUMNS = {'Інвентарний', 'Сума', 'Додл. інформація', 'Додл. Інформація'}


def parse_decimal(val) -> Decimal | None:
    if val is None:
        return None
    s = str(val).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def find_header_row(ws):
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if cell.value and str(cell.value).strip() == '№':
                return cell.row
    return None


def build_col_map(ws, header_row: int) -> dict:
    col_map = {}
    for cell in ws[header_row]:
        raw = str(cell.value).strip() if cell.value else ''
        if raw in COLUMN_MAP:
            col_map[cell.column] = COLUMN_MAP[raw]
    return col_map


def main():
    args = sys.argv[1:]
    if not args or args[0] in ('-h', '--help'):
        print(__doc__)
        sys.exit(0)

    filepath = args[0]
    dry_run = '--dry-run' in args

    if dry_run:
        print('=== DRY RUN — нічого не записується ===\n')

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    if not header_row:
        print("ПОМИЛКА: не знайдено рядок заголовків (шукаю колонку '№')")
        sys.exit(1)

    col_map = build_col_map(ws, header_row)
    print(f"Знайдено колонки: {list(col_map.values())}")
    print(f"Дані починаються з рядка {header_row + 1}\n")

    db = SessionLocal()
    created = skipped = errors = 0

    try:
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            row_vals = {}
            for cell in row:
                if cell.column in col_map and cell.value is not None:
                    row_vals[col_map[cell.column]] = cell.value

            if not row_vals.get('number') or not row_vals.get('name'):
                continue

            number = str(row_vals['number']).strip()
            name   = str(row_vals['name']).strip()

            # Перевірка дублікату
            existing = db.query(Item).filter(Item.number == number).first()
            if existing:
                print(f"  SKIP  #{number:>6}  {name[:60]}  (вже існує)")
                skipped += 1
                continue

            item = Item(
                number           = number,
                name             = name,
                nomenclature_code= str(row_vals['nomenclature_code']).strip() if row_vals.get('nomenclature_code') else None,
                serial_number    = str(row_vals['serial_number']).strip()     if row_vals.get('serial_number')    else None,
                unit_of_measure  = str(row_vals['unit_of_measure']).strip()   if row_vals.get('unit_of_measure')  else None,
                price            = parse_decimal(row_vals.get('price')),
                quantity         = parse_decimal(row_vals.get('quantity')) or Decimal('1'),
                item_type        = str(row_vals['item_type']).strip()          if row_vals.get('item_type')        else None,
                notes            = str(row_vals['notes']).strip()              if row_vals.get('notes')            else None,
                is_official      = True,
            )

            if dry_run:
                print(f"  DRY   #{number:>6}  {name[:60]}")
                created += 1
                continue

            try:
                db.add(item)
                db.commit()
                print(f"  OK    #{number:>6}  {name[:60]}")
                created += 1
            except Exception as e:
                db.rollback()
                print(f"  ERR   #{number:>6}  {name[:60]}  → {e}")
                errors += 1

    finally:
        db.close()

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Готово: {created} додано, {skipped} пропущено, {errors} помилок")


if __name__ == '__main__':
    main()
