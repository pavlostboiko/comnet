"""Admin-only inventory wipe + XLSX bulk import.

Wipe scope: items, asset_documents, item_documents, document_items,
movements, documents. Everything else (persons, recipients, users,
services, op_types, unit_settings) stays.

Import format is hard-coded against the layout the user already has —
the same column maps as the legacy scripts removed in 4f98167.
"""
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.auth import require_admin
from app.database import get_db
from app.models import (
    AssetDocument, Document, DocumentItem, Item, ItemDocument, ItemSplit,
    Movement, OpType, Person, Recipient, User,
)

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── Wipe ──────────────────────────────────────────────────────────────────

@router.post("/wipe-inventory", status_code=status.HTTP_200_OK)
def wipe_inventory(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    """Delete everything inventory-related. Persons/recipients/users
    /services/op_types/unit_settings survive."""
    # Order: leaf rows first, then dependents
    counts = {
        # Explicit delete for the counter — item_splits would CASCADE via
        # items, but we want to report the number that was removed.
        "item_splits":    db.query(ItemSplit).delete(synchronize_session=False),
        "document_items": db.query(DocumentItem).delete(synchronize_session=False),
        "item_documents": db.query(ItemDocument).delete(synchronize_session=False),
        "asset_documents": db.query(AssetDocument).delete(synchronize_session=False),
        "movements":      db.query(Movement).delete(synchronize_session=False),
        "documents":      db.query(Document).delete(synchronize_session=False),
        "items":          db.query(Item).delete(synchronize_session=False),
    }
    db.commit()
    return {"deleted": counts}


# ── Import helpers ────────────────────────────────────────────────────────

def _clean(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s or None


def _parse_decimal(val) -> Optional[Decimal]:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float, Decimal)):
        return Decimal(str(val))
    s = str(val).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_date(val) -> Optional[str]:
    """Return YYYY-MM-DD or original string. Movements.doc_date is VARCHAR."""
    if val is None or val == "":
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()


def _read_workbook(upload: UploadFile):
    blob = upload.file.read()
    try:
        return load_workbook(BytesIO(blob), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати XLSX: {e}")


# ── Items import ──────────────────────────────────────────────────────────

ITEMS_COLUMN_MAP = {
    "№":               "number",
    "Товар":           "name",
    "Код номер":       "nomenclature_code",
    "Серійний номер":  "serial_number",
    "Од. виміру":      "unit_of_measure",
    "Вартість":        "price",
    "Кіл-сть":         "quantity",
    "Категорія":       "item_type",
    "Де знаходиться":  "notes",
    "Видано":          "issued_recipient",  # resolved to recipient FK below
}

# Values in the «Видано» column that are NOT a recipient and must be skipped.
# Case-insensitive. Add more as the user reports them.
RECIPIENT_SKIP_VALUES = {"склад", ""}

# Placeholder tokens that stand in for «no serial number» in source files.
# Normalized to NULL on import so serial-vs-non-serial logic works correctly
# (Groups view, splits section visibility, dedup, etc.).
SERIAL_NONE_TOKENS = {"б/н", "бн", "б\\н", "н/д", "нд", "-", "—", "–", ""}


def _build_person_lookup(persons) -> dict:
    """Case-insensitive many-key → person.id map.

    Registers every reasonable spelling of a person so the movements import
    can match values like «Petro Ivanenko», «PETRO IVANENKO»,
    «Ivanenko Petro», or the existing search_name.
    """
    lookup: dict[str, int] = {}
    for p in persons:
        first = (p.first_name or "").strip().lower()
        last  = (p.last_name  or "").strip().lower()
        keys = set()
        if p.search_name:
            keys.add(p.search_name.strip().lower())
        if first and last:
            keys.add(f"{first} {last}")   # «petro ivanenko»
            keys.add(f"{last} {first}")   # «ivanenko petro»
        elif last:
            keys.add(last)
        elif first:
            keys.add(first)
        for k in keys:
            # First one wins if there's a collision — we simply skip later
            # persons with the same spelling. In practice search_name is
            # unique, and duplicate first+last is rare.
            lookup.setdefault(k, p.id)
    return lookup


def _resolve_person(raw: Optional[str], lookup: dict) -> Optional[int]:
    if not raw:
        return None
    key = raw.strip().lower()
    if not key:
        return None
    return lookup.get(key)


def _normalize_serial(val) -> Optional[str]:
    s = _clean(val)
    if s is None:
        return None
    if s.lower() in SERIAL_NONE_TOKENS:
        return None
    return s

TYPE_PREFIX_RE = re.compile(r"^\d+\.\s*")


def _find_items_header_row(ws) -> Optional[int]:
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if cell.value and str(cell.value).strip() == "№":
                return cell.row
    return None


def _build_items_col_map(ws, header_row: int) -> dict:
    col_map = {}
    for cell in ws[header_row]:
        raw = str(cell.value).strip() if cell.value else ""
        if raw in ITEMS_COLUMN_MAP:
            col_map[cell.column] = ITEMS_COLUMN_MAP[raw]
    return col_map


@router.post("/import/items", status_code=status.HTTP_200_OK)
def import_items(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    wb = _read_workbook(file)
    ws = wb.active

    header_row = _find_items_header_row(ws)
    if not header_row:
        raise HTTPException(400, "Не знайдено рядок заголовків (шукаю колонку '№')")
    col_map = _build_items_col_map(ws, header_row)
    if "number" not in col_map.values() or "name" not in col_map.values():
        raise HTTPException(400, "Обов'язкові колонки '№' і 'Товар' не знайдено")

    created, skipped, errors = 0, 0, []
    existing_numbers = set(n for (n,) in db.query(Item.number).all())
    # callsign_lc → recipient.id (case-insensitive lookup) — keep in-memory so
    # we don't hit the DB for every row; new ones append as we create them.
    recipients_by_callsign_lc = {
        (r.callsign or "").lower(): r.id for r in db.query(Recipient).all()
    }
    recipients_created = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        vals = {}
        for cell in row:
            if cell.column in col_map and cell.value is not None:
                vals[col_map[cell.column]] = cell.value
        if not vals.get("number") or not vals.get("name"):
            continue

        number = str(vals["number"]).strip()
        if number in existing_numbers:
            skipped += 1
            continue

        item_type = vals.get("item_type")
        if item_type:
            item_type = TYPE_PREFIX_RE.sub("", str(item_type)).strip()

        # Resolve «Видано» → recipient FK (find-or-create), with skip list.
        recipient_id = None
        raw_recipient = _clean(vals.get("issued_recipient"))
        if raw_recipient and raw_recipient.lower() not in RECIPIENT_SKIP_VALUES:
            key = raw_recipient.lower()
            if key in recipients_by_callsign_lc:
                recipient_id = recipients_by_callsign_lc[key]
            else:
                new_rec = Recipient(
                    callsign=raw_recipient, is_active=True, created_by=user.id,
                )
                db.add(new_rec)
                db.flush()
                recipients_by_callsign_lc[key] = new_rec.id
                recipient_id = new_rec.id
                recipients_created += 1

        try:
            item = Item(
                number            = number,
                name              = str(vals["name"]).strip(),
                nomenclature_code = _clean(vals.get("nomenclature_code")),
                serial_number     = _normalize_serial(vals.get("serial_number")),
                unit_of_measure   = _clean(vals.get("unit_of_measure")),
                price             = _parse_decimal(vals.get("price")),
                quantity          = _parse_decimal(vals.get("quantity")) or Decimal("1"),
                item_type         = item_type,
                notes             = _clean(vals.get("notes")),
                is_official       = True,
                issued_to_recipient_id = recipient_id,
                created_by        = user.id,
            )
            db.add(item)
            db.flush()
            # Same journaling path as /api/items POST — mirrors serial assignment
            # into the item_splits ledger so imported items show up in /history.
            from app.routers.items import _journal_serial_change
            _journal_serial_change(db, item, None, recipient_id, user.id)
            existing_numbers.add(number)
            created += 1
        except Exception as e:
            errors.append(f"#{number}: {e}")

    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "recipients_created": recipients_created,
    }


# ── Movements import ──────────────────────────────────────────────────────

# Column layout from the legacy import_movements.py — header in row 2,
# data starts row 3, zero-indexed positions:
MV_COLS = {
    "entry_date":        0,    # A
    "item_name":         1,    # B
    "unit_of_measure":   3,    # D
    "category":          4,    # E
    "qty_in":            5,    # F
    "qty_out":           6,    # G
    "from_unit":         7,    # H
    "to_unit":           8,    # I
    "mvo_from_name":     9,    # J
    "mvo_to_name":      10,    # K
    "basis":            11,    # L
    "doc_date":         12,    # M
    "doc_number":       13,    # N
    "serial_number":    15,    # P
    "nomenclature_code":17,    # R
    "price":            19,    # T
    "service":          20,    # U
    "item_card_num":    23,    # X
    "doc_type_excel":   24,    # Y
    "op_type":          25,    # Z
    "recipient_category":26,   # AA
}

OP_TYPE_MAP = {
    "надходження":           ("надходження", "акт"),
    "внутрішнє переміщення": ("переміщення", "накладна"),
    "переміщення":           ("переміщення", "накладна"),
}


def _col(row, idx):
    return row[idx].value if idx < len(row) else None


@router.post("/import/movements", status_code=status.HTTP_200_OK)
def import_movements(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    wb = _read_workbook(file)
    ws = wb.active

    # Persons lookup — case-insensitive, matches «first last», «last first»,
    # search_name, or just last name. See _build_person_lookup.
    persons_by_name = _build_person_lookup(db.query(Person).all())
    items_card_nums = set(n for (n,) in db.query(Item.number).all())

    created, skipped, errors, unmatched_persons = 0, 0, [], set()
    orphan_card_nums: set[str] = set()   # values in file's X column that
                                         # don't match any items.number
    auto_balanced: list[dict] = []       # rows we auto-mirrored qty_in↔qty_out
                                         # for internal transfers

    # Pass 1: collect document keys, create / dedupe Document rows
    doc_id_map = {}  # (op, form, doc_number, doc_date) → document.id
    rows_buffer = []

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
        if not _col(row, MV_COLS["entry_date"]) and not _col(row, MV_COLS["item_name"]):
            continue
        rows_buffer.append((i, row))

        op_raw = _clean(_col(row, MV_COLS["op_type"]))
        if not op_raw:
            continue
        op_pair = OP_TYPE_MAP.get(op_raw.lower())
        if not op_pair:
            continue
        op, form = op_pair
        doc_num  = _clean(_col(row, MV_COLS["doc_number"]))
        doc_date = _parse_date(_col(row, MV_COLS["doc_date"]))
        key = (op, form, doc_num, doc_date)
        if key in doc_id_map or (not doc_num and not doc_date):
            continue
        # Re-use existing matching document if any
        doc = db.query(Document).filter(
            Document.operation == op,
            Document.form == form,
            Document.doc_number == doc_num,
            Document.doc_date == doc_date,
        ).first()
        if not doc:
            doc = Document(
                operation  = op,
                form       = form,
                doc_number = doc_num,
                doc_date   = doc_date,
                from_unit  = _clean(_col(row, MV_COLS["from_unit"])),
                to_unit    = _clean(_col(row, MV_COLS["to_unit"])),
                basis      = _clean(_col(row, MV_COLS["basis"])),
                service    = _clean(_col(row, MV_COLS["service"])),
                status     = "signed",
                created_by = user.id,
            )
            db.add(doc)
            db.flush()
        doc_id_map[key] = doc.id

    # Pass 2: create movements
    for i, row in rows_buffer:
        try:
            op_raw = _clean(_col(row, MV_COLS["op_type"]))
            op_pair = OP_TYPE_MAP.get((op_raw or "").lower()) if op_raw else None
            op, form = op_pair if op_pair else (None, None)
            doc_num  = _clean(_col(row, MV_COLS["doc_number"]))
            doc_date = _parse_date(_col(row, MV_COLS["doc_date"]))
            doc_id   = doc_id_map.get((op, form, doc_num, doc_date)) if op else None

            mvo_from_name = _clean(_col(row, MV_COLS["mvo_from_name"]))
            mvo_to_name   = _clean(_col(row, MV_COLS["mvo_to_name"]))
            mvo_from_id = _resolve_person(mvo_from_name, persons_by_name)
            mvo_to_id   = _resolve_person(mvo_to_name,   persons_by_name)
            if mvo_from_name and not mvo_from_id:
                unmatched_persons.add(mvo_from_name)
            if mvo_to_name and not mvo_to_id:
                unmatched_persons.add(mvo_to_name)

            card_num = _clean(_col(row, MV_COLS["item_card_num"]))
            if card_num and card_num not in items_card_nums:
                orphan_card_nums.add(card_num)
            item_card_num = card_num if card_num in items_card_nums else None

            qty_in  = _parse_decimal(_col(row, MV_COLS["qty_in"]))
            qty_out = _parse_decimal(_col(row, MV_COLS["qty_out"]))
            # Auto-balance internal transfers: file conventions have only one
            # of qty_in/qty_out set for a внутрішнє переміщення row, but the
            # balance formula per (unit) is SUM(in) − SUM(out), so we need
            # BOTH sides set for the receiving unit to show a positive balance.
            # Mirror only when op is «переміщення» and exactly one side is set.
            if op == "переміщення":
                if qty_in and not qty_out:
                    qty_out = qty_in
                    auto_balanced.append({"row": i, "item": _clean(_col(row, MV_COLS["item_name"])), "qty": str(qty_in), "mirrored": "in→out"})
                elif qty_out and not qty_in:
                    qty_in = qty_out
                    auto_balanced.append({"row": i, "item": _clean(_col(row, MV_COLS["item_name"])), "qty": str(qty_out), "mirrored": "out→in"})

            db.add(Movement(
                document_id       = doc_id,
                entry_date        = _parse_date(_col(row, MV_COLS["entry_date"])),
                item_name         = _clean(_col(row, MV_COLS["item_name"])),
                item_card_num     = item_card_num,
                unit_of_measure   = _clean(_col(row, MV_COLS["unit_of_measure"])),
                category          = _clean(_col(row, MV_COLS["category"])),
                qty_in            = qty_in,
                qty_out           = qty_out,
                from_unit         = _clean(_col(row, MV_COLS["from_unit"])),
                to_unit           = _clean(_col(row, MV_COLS["to_unit"])),
                mvo_from_id       = mvo_from_id,
                mvo_to_id         = mvo_to_id,
                basis             = _clean(_col(row, MV_COLS["basis"])),
                doc_date          = doc_date,
                doc_number        = doc_num,
                serial_number     = _normalize_serial(_col(row, MV_COLS["serial_number"])),
                nomenclature_code = _clean(_col(row, MV_COLS["nomenclature_code"])),
                price             = _parse_decimal(_col(row, MV_COLS["price"])),
                service           = _clean(_col(row, MV_COLS["service"])),
                doc_type          = _clean(_col(row, MV_COLS["doc_type_excel"])),
                recipient_category = _clean(_col(row, MV_COLS["recipient_category"])),
                created_by        = user.id,
            ))
            created += 1
        except Exception as e:
            errors.append(f"row {i}: {e}")

    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "unmatched_persons": sorted(unmatched_persons),
        "orphan_card_nums": sorted(orphan_card_nums),
        "auto_balanced": auto_balanced,
        "auto_balanced_count": len(auto_balanced),
        "documents_created": len(doc_id_map),
    }


# ── Merge duplicates for non-serial items ────────────────────────────────

def _merge_key(item: Item) -> tuple:
    """Group items by (name, price, category, unit_of_measure). Serial items
    (with a serial_number) are excluded from merging by the caller."""
    return (
        (item.name or "").strip(),
        str(item.price) if item.price is not None else "",
        (item.category or "").strip(),
        (item.unit_of_measure or "").strip(),
    )


def _natural_key(number: str) -> tuple:
    """Sort key so pure-numeric numbers sort by value ('2' < '10')."""
    if number and number.isdigit():
        return (0, int(number))
    return (1, number or "")


def _group_nonserial_duplicates(db: Session) -> list[dict]:
    """Return list of merge candidates. Each group has ≥2 non-serial items
    sharing (name, price, category, unit)."""
    items = db.query(Item).filter(Item.serial_number.is_(None)).all()
    buckets: dict[tuple, list[Item]] = {}
    for it in items:
        buckets.setdefault(_merge_key(it), []).append(it)

    groups = []
    for key, cards in buckets.items():
        if len(cards) < 2:
            continue
        cards.sort(key=lambda c: _natural_key(c.number))
        winner = cards[0]
        losers = cards[1:]
        groups.append({
            "key": {
                "name": key[0],
                "price": key[1],
                "category": key[2],
                "unit_of_measure": key[3],
            },
            "winner_id": winner.id,
            "winner_number": winner.number,
            "loser_ids": [c.id for c in losers],
            "loser_numbers": [c.number for c in losers],
            "cards_count": len(cards),
            "total_quantity": str(sum((Decimal(c.quantity or 0) for c in cards), Decimal(0))),
        })
    # Sort by name for a stable preview
    groups.sort(key=lambda g: g["key"]["name"])
    return groups


@router.get("/merge-nonserial-duplicates/preview")
def merge_nonserial_preview(
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    groups = _group_nonserial_duplicates(db)
    total_cards_to_remove = sum(len(g["loser_ids"]) for g in groups)
    return {
        "groups_found": len(groups),
        "cards_to_remove": total_cards_to_remove,
        "groups": groups[:100],  # preview cap
    }


@router.post("/merge-nonserial-duplicates")
def merge_nonserial_apply(
    db: Session = Depends(get_db),
    _: User = Depends(require_admin),
):
    groups = _group_nonserial_duplicates(db)
    merged_groups = 0
    removed_cards = 0

    for group in groups:
        winner = db.get(Item, group["winner_id"])
        losers = [db.get(Item, lid) for lid in group["loser_ids"]]
        losers = [l for l in losers if l is not None]
        if not losers:
            continue

        # Sum quantity
        total = Decimal(winner.quantity or 0)
        for l in losers:
            total += Decimal(l.quantity or 0)
        winner.quantity = total

        # Merge notes with `; `
        notes_parts = [winner.notes.strip()] if winner.notes and winner.notes.strip() else []
        for l in losers:
            if l.notes and l.notes.strip() and l.notes.strip() not in notes_parts:
                notes_parts.append(l.notes.strip())
        if notes_parts:
            winner.notes = "; ".join(notes_parts)

        # Merge issued_to: keep winner's if set, else first non-null loser
        if winner.issued_to_recipient_id is None:
            for l in losers:
                if l.issued_to_recipient_id is not None:
                    winner.issued_to_recipient_id = l.issued_to_recipient_id
                    break

        # Reassign related rows
        loser_ids = [l.id for l in losers]
        loser_numbers = [l.number for l in losers]

        # movements.item_card_num → winner.number (references items.number)
        db.query(Movement).filter(Movement.item_card_num.in_(loser_numbers)).update(
            {"item_card_num": winner.number}, synchronize_session=False,
        )
        # document_items.item_id → winner.id
        db.query(DocumentItem).filter(DocumentItem.item_id.in_(loser_ids)).update(
            {"item_id": winner.id}, synchronize_session=False,
        )
        # item_splits.item_id → winner.id
        db.query(ItemSplit).filter(ItemSplit.item_id.in_(loser_ids)).update(
            {"item_id": winner.id}, synchronize_session=False,
        )
        # item_documents links: move to winner if not already linked, else drop
        existing_doc_ids = {
            r.doc_id for r in db.query(ItemDocument).filter(ItemDocument.item_id == winner.id).all()
        }
        for lid in loser_ids:
            links = db.query(ItemDocument).filter(ItemDocument.item_id == lid).all()
            for lnk in links:
                if lnk.doc_id in existing_doc_ids:
                    db.delete(lnk)
                else:
                    lnk.item_id = winner.id
                    existing_doc_ids.add(lnk.doc_id)

        # Now safe to delete losers
        for l in losers:
            db.delete(l)
        removed_cards += len(losers)
        merged_groups += 1

    db.commit()
    return {"merged_groups": merged_groups, "removed_cards": removed_cards}
