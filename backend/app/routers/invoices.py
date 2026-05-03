import io
from decimal import Decimal
from typing import List

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.database import get_db
from app.models import Person, PrintDocument, PrintDocumentItem, UnitSettings
from app.schemas import InvoiceCreate, InvoiceListRead, InvoiceRead, InvoiceUpdate

router = APIRouter(prefix="/api/invoices", tags=["invoices"])

EXTRA_FIELDS = [
    "validity_date", "composed_date", "composed_location", "operation_date",
    "service", "op_type_text", "basis", "responsible_recipient",
    "sender_id", "receiver_id", "commander_id", "mvo_from_id", "mvo_to_id",
    "total_qty_words", "total_amount_words",
]


def _doc_to_read(doc: PrintDocument) -> dict:
    extra = doc.extra_data or {}
    items_sorted = sorted(doc.items, key=lambda x: (x.sort_order or 0, x.id))
    return {
        "id": doc.id,
        "doc_number": doc.doc_number,
        "doc_date": doc.doc_date,
        "from_unit": doc.from_unit,
        "to_unit": doc.to_unit,
        "status": doc.status,
        **{f: extra.get(f) for f in EXTRA_FIELDS},
        "items": [
            {
                "id": it.id,
                "sort_order": it.sort_order,
                "item_name": it.item_name,
                "nomenclature_code": it.nomenclature_code,
                "unit_of_measure": it.unit_of_measure,
                "category": it.category,
                "quantity": it.quantity,
                "qty_received": it.qty_received,
                "price": it.price,
                "amount": it.amount,
                "notes": it.notes,
            }
            for it in items_sorted
        ],
    }


@router.get("", response_model=List[InvoiceListRead])
def list_invoices(db: Session = Depends(get_db), _=Depends(get_current_user)):
    docs = (
        db.query(PrintDocument)
        .filter(PrintDocument.doc_type == "накладна_25")
        .order_by(PrintDocument.created_at.desc())
        .all()
    )
    return docs


@router.get("/{doc_id}")
def get_invoice(doc_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    doc = db.query(PrintDocument).filter(
        PrintDocument.id == doc_id,
        PrintDocument.doc_type == "накладна_25",
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return _doc_to_read(doc)


@router.post("", status_code=status.HTTP_201_CREATED)
def create_invoice(
    payload: InvoiceCreate,
    db: Session = Depends(get_db),
    user=Depends(get_current_user),
):
    extra = {f: getattr(payload, f) for f in EXTRA_FIELDS}
    doc = PrintDocument(
        doc_type="накладна_25",
        doc_number=payload.doc_number,
        doc_date=payload.doc_date,
        from_unit=payload.from_unit,
        to_unit=payload.to_unit,
        status=payload.status,
        extra_data=extra,
        created_by=user.id,
    )
    db.add(doc)
    db.flush()

    for idx, it in enumerate(payload.items):
        row = PrintDocumentItem(
            document_id=doc.id,
            sort_order=it.sort_order if it.sort_order is not None else idx,
            item_name=it.item_name,
            nomenclature_code=it.nomenclature_code,
            unit_of_measure=it.unit_of_measure,
            category=it.category,
            quantity=it.quantity,
            qty_received=it.qty_received,
            price=it.price,
            amount=it.amount,
            notes=it.notes,
        )
        db.add(row)

    db.commit()
    db.refresh(doc)
    return _doc_to_read(doc)


@router.put("/{doc_id}")
def update_invoice(
    doc_id: int,
    payload: InvoiceUpdate,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    doc = db.query(PrintDocument).filter(
        PrintDocument.id == doc_id,
        PrintDocument.doc_type == "накладна_25",
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")

    doc.doc_number = payload.doc_number
    doc.doc_date = payload.doc_date
    doc.from_unit = payload.from_unit
    doc.to_unit = payload.to_unit
    doc.status = payload.status
    doc.extra_data = {f: getattr(payload, f) for f in EXTRA_FIELDS}

    # replace items
    for it in list(doc.items):
        db.delete(it)
    db.flush()

    for idx, it in enumerate(payload.items):
        row = PrintDocumentItem(
            document_id=doc.id,
            sort_order=it.sort_order if it.sort_order is not None else idx,
            item_name=it.item_name,
            nomenclature_code=it.nomenclature_code,
            unit_of_measure=it.unit_of_measure,
            category=it.category,
            quantity=it.quantity,
            qty_received=it.qty_received,
            price=it.price,
            amount=it.amount,
            notes=it.notes,
        )
        db.add(row)

    db.commit()
    db.refresh(doc)
    return _doc_to_read(doc)


@router.delete("/{doc_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_invoice(doc_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    doc = db.query(PrintDocument).filter(
        PrintDocument.id == doc_id,
        PrintDocument.doc_type == "накладна_25",
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(doc)
    db.commit()


@router.get("/{doc_id}/export/xlsx")
def export_invoice_xlsx(
    doc_id: int,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    doc = db.query(PrintDocument).filter(
        PrintDocument.id == doc_id,
        PrintDocument.doc_type == "накладна_25",
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")

    extra = doc.extra_data or {}
    settings = db.query(UnitSettings).first()

    def person_str(pid, genitive=False):
        if not pid:
            return ""
        p = db.query(Person).filter(Person.id == pid).first()
        if not p:
            return ""
        parts = []
        if genitive:
            if p.rank_genitive:
                parts.append(p.rank_genitive)
            name_parts = [p.last_name_genitive or p.last_name or ""]
            if p.first_name:
                name_parts.append(p.first_name[0] + ".")
            if p.patronymic:
                name_parts.append(p.patronymic[0] + ".")
            parts.append(" ".join(name_parts).strip())
        else:
            if p.rank:
                parts.append(p.rank)
            name_parts = [p.last_name or ""]
            if p.first_name:
                name_parts.append(p.first_name[0] + ".")
            if p.patronymic:
                name_parts.append(p.patronymic[0] + ".")
            parts.append(" ".join(name_parts).strip())
        return " ".join(parts).strip()

    def person_position(pid, genitive=False):
        if not pid:
            return ""
        p = db.query(Person).filter(Person.id == pid).first()
        if not p:
            return ""
        return (p.position_genitive if genitive else p.position) or ""

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Накладна"

    thin = Side(style="thin")
    thick = Side(style="medium")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick = Border(left=thick, right=thick, top=thick, bottom=thick)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    bold = Font(bold=True)

    def cell(row, col, value="", font=None, align=None, border=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if align:
            c.alignment = align
        if border:
            c.border = border
        return c

    def merge(r1, c1, r2, c2, value="", font=None, align=None, border=None):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        c = ws.cell(row=r1, column=c1, value=value)
        if font:
            c.font = font
        if align:
            c.alignment = align
        if border:
            c.border = border
        return c

    # Column widths (A–J = 10 cols for table, plus helpers)
    col_widths = [5, 30, 18, 10, 10, 14, 12, 12, 14, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    unit_name = (settings.name or "") if settings else ""
    edrpou = (settings.edrpou or "") if settings else ""

    r = 1
    # Row 1-2: unit header + validity
    merge(r, 1, r+1, 6, unit_name, font=bold, align=left)
    merge(r, 7, r, 10, f"Термін дії накладної: {extra.get('validity_date','')}", align=left)
    ws.row_dimensions[r].height = 20
    r += 1
    merge(r, 7, r, 10, f"ЄДРПОУ: {edrpou}", align=left)
    ws.row_dimensions[r].height = 15
    r += 1

    # Row 3: НАКЛАДНА title
    merge(r, 1, r, 10, f"НАКЛАДНА (ВИМОГА) № {doc.doc_number or ''}",
          font=Font(bold=True, size=14), align=center)
    ws.row_dimensions[r].height = 25
    r += 1

    # Row 4: location + date
    location = extra.get("composed_location") or (settings.location if settings else "") or ""
    merge(r, 1, r, 5, f"Місце складання: {location}", align=left)
    merge(r, 6, r, 10, f"Дата: {extra.get('composed_date','')}", align=left)
    ws.row_dimensions[r].height = 18
    r += 1

    # Row 5: operation date
    merge(r, 1, r, 10, f"Дата операцій: {extra.get('operation_date','')}", align=left)
    ws.row_dimensions[r].height = 15
    r += 1

    # Row 6: service + op_type
    merge(r, 1, r, 5, f"Служба: {extra.get('service','')}", align=left)
    merge(r, 6, r, 10, f"Вид операції: {extra.get('op_type_text','')}", align=left)
    ws.row_dimensions[r].height = 15
    r += 1

    # Row 7: basis
    merge(r, 1, r, 10, f"Підстава: {extra.get('basis','')}", align=left)
    ws.row_dimensions[r].height = 15
    r += 1

    # Row 8: from/to
    merge(r, 1, r, 5, f"Звідки: {doc.from_unit or ''}", align=left)
    merge(r, 6, r, 10, f"Куди: {doc.to_unit or ''}", align=left)
    ws.row_dimensions[r].height = 15
    r += 1

    # Row 9: sender/receiver
    sender_pos = person_position(extra.get("sender_id"))
    sender_name = person_str(extra.get("sender_id"))
    receiver_pos = person_position(extra.get("receiver_id"))
    receiver_name = person_str(extra.get("receiver_id"))
    merge(r, 1, r, 5, f"Передає: {sender_pos} {sender_name}", align=left)
    merge(r, 6, r, 10, f"Приймає: {receiver_pos} {receiver_name}", align=left)
    ws.row_dimensions[r].height = 18
    r += 1

    # Responsible recipient
    merge(r, 1, r, 10, f"Відповідальна особа-отримувач: {extra.get('responsible_recipient','')}", align=left)
    ws.row_dimensions[r].height = 15
    r += 1

    header_start = r
    # Table header
    headers = [
        "№\nз/п", "Назва майна або однорідна група",
        "Код номен-клатури", "Одиниця виміру", "Категорія\n(сорт)",
        "Вартість за одиницю", "Кількість\nвідправлено",
        "Кількість\nприйнято", "Сума", "Примітка",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=col, value=h)
        c.font = bold
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[r].height = 40
    r += 1

    items_sorted = sorted(doc.items, key=lambda x: (x.sort_order or 0, x.id))
    total_qty = Decimal(0)
    total_amount = Decimal(0)

    for idx, it in enumerate(items_sorted, 1):
        row_data = [
            idx,
            it.item_name or "",
            it.nomenclature_code or "",
            it.unit_of_measure or "",
            it.category or "",
            float(it.price) if it.price else "",
            float(it.quantity) if it.quantity else "",
            float(it.qty_received) if it.qty_received else "",
            float(it.amount) if it.amount else "",
            it.notes or "",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border_all
            c.alignment = center if col in (1, 3, 4, 5, 6, 7, 8, 9) else left
        ws.row_dimensions[r].height = 20
        if it.quantity:
            total_qty += Decimal(str(it.quantity))
        if it.amount:
            total_amount += Decimal(str(it.amount))
        r += 1

    # Footer row: totals
    merge(r, 1, r, 5, "Разом:", font=bold, align=left)
    ws.cell(row=r, column=7, value=float(total_qty)).border = border_all
    ws.cell(row=r, column=9, value=float(total_amount)).border = border_all
    for col in [1, 2, 3, 4, 5, 6, 8, 10]:
        ws.cell(row=r, column=col).border = border_all
    ws.row_dimensions[r].height = 18
    r += 1

    # Words
    r += 1
    cmd_pos = person_position(extra.get("commander_id"))
    cmd_name = person_str(extra.get("commander_id"))
    merge(r, 1, r, 4, f"Керівник: {cmd_pos}", align=left)
    merge(r, 5, r, 8, "підпис ____________", align=left)
    merge(r, 9, r, 10, cmd_name, align=left)
    ws.row_dimensions[r].height = 18
    r += 1

    r += 1
    merge(r, 1, r, 10,
          f"Всього передано: {extra.get('total_qty_words','')} одиниць, на суму {extra.get('total_amount_words','')} грн.",
          align=left)
    ws.row_dimensions[r].height = 18
    r += 2

    # Page 2 header
    merge(r, 1, r, 10, "ЗВОРОТНІЙ БІК", font=bold, align=center)
    ws.row_dimensions[r].height = 20
    r += 2

    mvo_from_pos = person_position(extra.get("mvo_from_id"))
    mvo_from_name = person_str(extra.get("mvo_from_id"))
    mvo_to_pos = person_position(extra.get("mvo_to_id"))
    mvo_to_name = person_str(extra.get("mvo_to_id"))

    merge(r, 1, r, 5, f"МВО здав: {mvo_from_pos}", align=left)
    merge(r, 6, r, 8, "підпис ____________", align=left)
    merge(r, 9, r, 10, mvo_from_name, align=left)
    ws.row_dimensions[r].height = 18
    r += 2

    merge(r, 1, r, 5, f"МВО прийняв: {mvo_to_pos}", align=left)
    merge(r, 6, r, 8, "підпис ____________", align=left)
    merge(r, 9, r, 10, mvo_to_name, align=left)
    ws.row_dimensions[r].height = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"nakладна_{doc.doc_number or doc.id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
