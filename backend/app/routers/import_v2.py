"""v2 Items import (Фаза 6).

Один файл Items → розкладається на v2-сутності (find-or-create по ходу):
  рядок → служба + номенклатура (+ instance якщо серійне)
        → склад підрозділу (з колонки «Де») → receipt-рух
        → якщо в «Де» є людина → assignment на неї.

Колонка «Де» = «<підрозділ> [людина]». Підрозділ визначається найдовшим
збігом з ІСНУЮЧИМИ підрозділами (тому їх варто завести до імпорту); решта = ПІБ.
is_official за замовчуванням True (державне).
"""
from datetime import date as date_cls, datetime
from decimal import Decimal
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.auth import require_admin
from app.database import get_db
from app.custody_snapshot import doc_sort_key, snap_nakladna
from app.models import (
    Assignment, CustodyDocument, CustodyMovement, Instance, Mvo, Nomenclature,
    Person, Service, Unit, User, Warehouse,
)
from app.routers.admin import (
    _build_person_lookup, _clean, _normalize_serial, _parse_decimal, _parse_date,
    _resolve_person,
)
from app.routers.assignments import _active_issued
from app.routers.custody import balance_of
from app.routers.structure import ensure_warehouse_for_service, ensure_warehouse_for_unit

router = APIRouter(prefix="/api/admin/v2", tags=["admin-v2"])

# Значення «Де», що означають «склад служби» (не підрозділ, не видано особі).
WAREHOUSE_TOKENS = {"склад", "на складі", "склад служби"}


def is_service_warehouse_location(raw) -> bool:
    return (str(raw).strip().lower() if raw is not None else "") in WAREHOUSE_TOKENS

# header → logical key
COLS = {
    "№": "card_number",
    "Назва": "name", "Товар": "name",
    "Служба": "service",
    "Категорія": "category",
    "Серійний номер": "serial",
    "Од. виміру": "uom", "Од.виміру": "uom",
    "Вартість": "price",
    "Кіл-сть": "qty", "Кількість": "qty",
    "Код номер": "code", "Код": "code",
    "Де": "location", "Де знаходиться": "location",
    "Дата видачі": "issued_at",
}


def _find_header_row(ws) -> Optional[int]:
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            v = str(cell.value).strip() if cell.value else ""
            if v in ("Назва", "Товар"):
                return cell.row
    return None


def _col_map(ws, header_row: int) -> dict:
    m = {}
    for cell in ws[header_row]:
        raw = str(cell.value).strip() if cell.value else ""
        if raw in COLS:
            m[cell.column] = COLS[raw]
    return m


@router.post("/wipe", status_code=status.HTTP_200_OK)
def wipe_v2_inventory(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    """Скидає v2 інвентарний шар для повторного імпорту: видачі, рухи,
    документи, екземпляри, номенклатуру. Довідники (служби/підрозділи/склади/
    особи) лишаються — імпорт їх find-or-create. Документи чистимо теж, бо
    backfill створює їх під час імпорту «Переміщення» — інакше накопичуються."""
    counts = {
        "assignments": db.query(Assignment).delete(synchronize_session=False),
        "custody_movements": db.query(CustodyMovement).delete(synchronize_session=False),
        "custody_documents": db.query(CustodyDocument).delete(synchronize_session=False),
        "instances": db.query(Instance).delete(synchronize_session=False),
        "nomenclature": db.query(Nomenclature).delete(synchronize_session=False),
    }
    db.commit()
    return {"deleted": counts}


@router.post("/wipe-persons", status_code=status.HTTP_200_OK)
def wipe_persons_v2(db: Session = Depends(get_db), _: User = Depends(require_admin)):
    """Очистити список осіб (для повного переналаштування перед реімпортом).
    Особи, прив'язані до облікового запису (МВО-логіну), НЕ видаляються — інакше
    логін втратить auto-scope на підрозділ. Якщо є активні видачі — блокуємо
    (спершу «Очистити інвентар»), щоб не втратити видачі каскадом разом з особами."""
    if db.query(Assignment).count() > 0:
        raise HTTPException(
            400, "Спершу очистіть інвентар — у осіб є видачі (assignments)")
    # Зберігаємо осіб, прив'язаних до логіну (МВО-scope) І осіб у журналі МВО
    # (інакше каскад знесе журнал і його довелось би вносити заново).
    linked_ids = {pid for (pid,) in
                  db.query(User.person_id).filter(User.person_id.isnot(None)).all()}
    mvo_ids = {pid for (pid,) in
               db.query(Mvo.person_id).filter(Mvo.person_id.isnot(None)).distinct()}
    keep = linked_ids | mvo_ids
    q = db.query(Person)
    if keep:
        q = q.filter(~Person.id.in_(keep))
    deleted = q.delete(synchronize_session=False)
    db.commit()
    return {"deleted_persons": deleted, "kept_linked_to_users": len(linked_ids),
            "kept_mvo": len(mvo_ids)}


PERSON_COLS = {
    "ПІБ": "fullname", "П.І.Б": "fullname", "П.І.Б.": "fullname", "ПiБ": "fullname",
    "ІПН": "ipn", "IПН": "ipn", "Ідентифікаційний код": "ipn",
    "Позивний": "callsign",
}


@router.post("/import/persons", status_code=status.HTTP_200_OK)
def import_persons_v2(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Імпорт осіб: колонки ПІБ (одна, розділення пробілом) / ІПН / Позивний.
    Ключ — ІПН: є в базі → оновити (ПІБ, позивний, active); нема → створити;
    ІПН у базі, якого нема у файлі → деактивувати (is_active=False). Рядок без
    ІПН пропускаємо. `search_name` = повний ПІБ (для матчингу МВО/рухів по J/K)."""
    try:
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати XLSX: {e}")
    ws = wb.active

    header_row = None
    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if str(cell.value or "").strip() in ("ПІБ", "П.І.Б", "П.І.Б.", "ПiБ"):
                header_row = cell.row
                break
        if header_row:
            break
    if not header_row:
        raise HTTPException(400, "Не знайдено заголовок «ПІБ»")
    col = {}
    for cell in ws[header_row]:
        key = PERSON_COLS.get(str(cell.value or "").strip())
        if key:
            col[key] = cell.column
    if "ipn" not in col:
        raise HTTPException(400, "Не знайдено колонку «ІПН»")

    def cv(cells, key):
        c = col.get(key)
        return cells[c - 1].value if c and c - 1 < len(cells) else None

    counts = {"rows": 0, "created": 0, "updated": 0, "deactivated": 0, "skipped": 0}
    errors = []
    existing = {p.ipn: p for p in db.query(Person).filter(Person.ipn.isnot(None)).all() if p.ipn}
    seen = set()

    for cells in ws.iter_rows(min_row=header_row + 1, values_only=False):
        fullname = _clean(cv(cells, "fullname"))
        if not fullname:
            continue
        counts["rows"] += 1
        ipn = _clean(cv(cells, "ipn"))
        if not ipn:
            counts["skipped"] += 1
            errors.append(f"«{fullname}»: без ІПН — пропущено")
            continue
        parts = fullname.split()
        last = parts[0] if parts else None
        first = parts[1] if len(parts) > 1 else None
        patr = " ".join(parts[2:]) if len(parts) > 2 else None
        callsign = _clean(cv(cells, "callsign"))
        seen.add(ipn)
        p = existing.get(ipn)
        if p:
            p.last_name, p.first_name, p.patronymic = last, first, patr
            p.callsign, p.is_active, p.search_name = callsign, True, fullname.lower()
            counts["updated"] += 1
        else:
            db.add(Person(last_name=last, first_name=first, patronymic=patr, ipn=ipn,
                          callsign=callsign, is_active=True, search_name=fullname.lower()))
            counts["created"] += 1

    # Особи з ІПН, яких нема у файлі → деактивувати (без ІПН — не чіпаємо).
    for p in existing.values():
        if p.ipn not in seen and p.is_active:
            p.is_active = False
            counts["deactivated"] += 1

    db.commit()
    return {**counts, "errors": errors[:100]}


@router.post("/import/items", status_code=status.HTTP_200_OK)
def import_items_v2(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    try:
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати XLSX: {e}")
    ws = wb.active
    header_row = _find_header_row(ws)
    if not header_row:
        raise HTTPException(400, "Не знайдено заголовок «Назва»/«Товар»")
    cmap = _col_map(ws, header_row)

    # Items = КАТАЛОГ: лише номенклатура + серійні екземпляри. Розміщення на
    # складах (баланси) робить окремий імпорт Переміщень (рішення 2026-07-24),
    # щоб не подвоювати. Колонка «Де» тут ігнорується.
    services = {s.name.strip().lower(): s for s in db.query(Service).all()}
    noms = {(n.name.strip().lower(), n.service_id): n for n in db.query(Nomenclature).all()}

    counts = {"rows": 0, "nomenclature": 0, "instances": 0, "services_created": 0}
    errors = []

    def get_service(name):
        key = name.strip().lower()
        s = services.get(key)
        if not s:
            s = Service(name=name.strip())
            db.add(s); db.flush()
            ensure_warehouse_for_service(db, s)  # склад служби знадобиться Переміщенням
            services[key] = s
            counts["services_created"] += 1
        return s

    def get_nomenclature(name, service, is_serial, uom, price, code, category):
        key = (name.strip().lower(), service.id)
        n = noms.get(key)
        if not n:
            n = Nomenclature(name=name.strip(), service_id=service.id, category=category,
                             is_serialized=is_serial, unit_of_measure=uom, price=price, code=code)
            db.add(n); db.flush()
            noms[key] = n
            counts["nomenclature"] += 1
        return n

    def col(row, key):
        for c, k in cmap.items():
            if k == key:
                return row[c - 1].value
        return None

    for row in ws.iter_rows(min_row=header_row + 1):
        name = _clean(col(row, "name"))
        if not name:
            continue
        counts["rows"] += 1
        try:
            svc_name = _clean(col(row, "service"))
            if not svc_name:
                errors.append(f"«{name}»: порожня служба — пропущено")
                continue
            service = get_service(svc_name)
            serial = _normalize_serial(col(row, "serial"))
            is_serial = serial is not None
            nom = get_nomenclature(
                name, service, is_serial,
                _clean(col(row, "uom")), _parse_decimal(col(row, "price")), _clean(col(row, "code")),
                _clean(col(row, "category")),
            )
            # серійний екземпляр — без розміщення (склад заповнить Переміщення)
            if is_serial:
                inst = db.query(Instance).filter(Instance.serial_no == serial).first()
                card = _clean(col(row, "card_number"))
                if not inst:
                    db.add(Instance(nomenclature_id=nom.id, serial_no=serial,
                                    card_number=card, is_official=nom.is_official))
                    counts["instances"] += 1
                elif card and not inst.card_number:
                    inst.card_number = card
        except Exception as e:
            errors.append(f"«{name}»: {e}")

    db.commit()
    return {**counts, "errors": errors}


# ── Переміщення (custody movements) — жорсткий layout v1 ─────────────────────

MV_COLS = {
    "entry_date": 0, "item_name": 1, "unit_of_measure": 3, "qty_in": 5,
    "qty_out": 6, "from_unit": 7, "to_unit": 8, "doc_date": 12, "doc_number": 13,
    "serial_number": 15, "price": 19, "service": 20, "card_number": 23,  # X = «Поле 12»
    "doc_type": 24, "operation": 25,   # Y = Тип документа, Z = Запасне поле2
    "mvo_from": 9, "mvo_to": 10,       # J = МВО звідки, K = МВО куди
}


def _op_from_hint(z) -> Optional[str]:
    """Тип операції з колонки Z (Запасне поле2): надходження → receipt,
    внутрішнє переміщення → transfer. None → невідомо (лишити евристику)."""
    s = (z or "").strip().lower()
    if "надходж" in s:
        return "receipt"
    if "переміщ" in s or "перемищ" in s:
        return "transfer"
    return None


def _form_from_hint(y) -> str:
    """Форма документа з колонки Y (Тип документа): «Акт …» → акт,
    інакше (в т.ч. «Накладна (вимога)») → накладна."""
    s = (y or "").strip().lower()
    return "акт" if s.startswith("акт") else "накладна"


def _build_mvo_journal(db: Session, mvo_obs: dict, counts: dict) -> None:
    """З колонок J/K будуємо історичний журнал МВО складів. Для кожного складу:
    сортуємо спостереження (дата, особа) за датою, зміна МВО закриває попередній
    період (to_date = день перед наступним), остання — активна. Журнал складу
    ПЕРЕбудовується з файлу (щоб реімпорт не плодив дублів; ручні записи інших
    складів не чіпаються)."""
    from datetime import timedelta
    for wh_id, obs in mvo_obs.items():
        if not obs:
            continue
        obs.sort(key=lambda x: x[0])
        periods = []                       # [from_date, person_id] у точках зміни
        for d, pid in obs:
            if not periods or periods[-1][1] != pid:
                periods.append([d, pid])
        db.query(Mvo).filter(Mvo.kind == "warehouse", Mvo.warehouse_id == wh_id) \
            .delete(synchronize_session=False)
        for idx, (fd, pid) in enumerate(periods):
            to_d = None
            if idx + 1 < len(periods):
                to_d = periods[idx + 1][0] - timedelta(days=1)
                if to_d < fd:
                    to_d = fd
            db.add(Mvo(kind="warehouse", warehouse_id=wh_id, person_id=pid,
                       from_date=fd, to_date=to_d))
            counts["mvo_created"] += 1


@router.post("/import/movements", status_code=status.HTTP_200_OK)
def import_movements_v2(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Файл Переміщень (той самий layout, що v1) → custody_movements.
    from_unit/to_unit резолвляться у склади: «склад»/назва служби → склад служби,
    інакше → склад підрозділу (find-or-create). Історичні дані НЕ валідуються
    (вставляються як є); поточне розташування серійного = to-склад руху, що
    хронологічно ПІЗНІШИЙ (дата → номер накладної → id), незалежно від порядку
    рядків у файлі."""
    try:
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати XLSX: {e}")
    ws = wb.active

    services = {s.name.strip().lower(): s for s in db.query(Service).all()}
    units = {u.name.strip().lower(): u for u in db.query(Unit).all()}
    noms = {(n.name.strip().lower(), n.service_id): n for n in db.query(Nomenclature).all()}

    counts = {"rows": 0, "movements": 0, "skipped": 0, "documents_created": 0,
              "services_created": 0, "units_created": 0, "instances_created": 0,
              # діагностика: скільки рядків класифіковано за колонкою Z, а скільки
              # впало у стару евристику (from/to) — щоб побачити розбіжність колонок.
              "classified_by_column": 0, "by_heuristic": 0, "mvo_created": 0}
    # Групування рухів у накладні по (номер, звідки, куди). Заповнюється у циклі.
    doc_groups: dict = {}
    # Рухи серійних екземплярів — для детермінованого розміщення в кінці імпорту
    # (не «останній рядок перемагає», а хронологічно пізніший рух). Ключ: instance.
    serial_movements: dict = {}
    # МВО з колонок J/K: warehouse_id → [(дата, person_id)]; журнал будуємо в кінці.
    mvo_obs: dict = {}
    person_lookup = _build_person_lookup(db.query(Person).all())
    mvo_unmatched: set = set()
    errors = []

    def get_service(name):
        key = name.strip().lower()
        s = services.get(key)
        if not s:
            s = Service(name=name.strip()); db.add(s); db.flush()
            ensure_warehouse_for_service(db, s)
            services[key] = s; counts["services_created"] += 1
        return s

    def get_unit(name):
        key = name.strip().lower()
        u = units.get(key)
        if not u:
            u = Unit(name=name.strip()); db.add(u); db.flush()
            ensure_warehouse_for_unit(db, u)
            units[key] = u; counts["units_created"] += 1
        return u

    def wh_of_service(s):
        return db.query(Warehouse).filter(Warehouse.service_id == s.id).first()

    def wh_of_unit(u):
        return db.query(Warehouse).filter(Warehouse.unit_id == u.id).first()

    def resolve_wh(name, service):
        s = (name or "").strip()
        if not s:
            return None  # зовні (надходження/списання)
        if is_service_warehouse_location(s):        # «склад» → склад цієї служби
            return wh_of_service(service)
        if s.lower() in services:                   # назва служби → її склад
            return wh_of_service(services[s.lower()])
        return wh_of_unit(get_unit(s))              # інакше — склад підрозділу

    def get_nomenclature(name, service, is_serial, uom, price):
        key = (name.strip().lower(), service.id)
        n = noms.get(key)
        if not n:
            n = Nomenclature(name=name.strip(), service_id=service.id,
                             is_serialized=is_serial, unit_of_measure=uom, price=price)
            db.add(n); db.flush(); noms[key] = n
        return n

    def mc(row, key):
        idx = MV_COLS[key]
        return row[idx].value if idx < len(row) else None

    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
        cells = row
        name = _clean(mc(cells, "item_name"))
        if not name:
            continue
        counts["rows"] += 1
        try:
            svc_name = _clean(mc(cells, "service"))
            if not svc_name:
                counts["skipped"] += 1
                errors.append(f"рядок {i} «{name}»: порожня служба")
                continue
            service = get_service(svc_name)
            serial = _normalize_serial(mc(cells, "serial_number"))
            is_serial_row = serial is not None
            uom = _clean(mc(cells, "unit_of_measure"))
            price = _parse_decimal(mc(cells, "price"))
            card = _clean(mc(cells, "card_number"))
            nom = get_nomenclature(name, service, is_serial_row, uom, price)

            qin = _parse_decimal(mc(cells, "qty_in")) or Decimal(0)
            qout = _parse_decimal(mc(cells, "qty_out")) or Decimal(0)
            raw_from = mc(cells, "from_unit")
            to_wh = resolve_wh(mc(cells, "to_unit"), service)
            form = _form_from_hint(mc(cells, "doc_type"))       # Y = Тип документа
            op_hint = _op_from_hint(mc(cells, "operation"))     # Z = Запасне поле2
            counterparty = None

            if op_hint == "receipt":                            # надходження ззовні
                mtype = "receipt"; from_wh = None
                counterparty = _clean(raw_from)                 # «від кого»
            elif op_hint == "transfer":                         # внутрішнє переміщення
                mtype = "transfer"; from_wh = resolve_wh(raw_from, service)
            else:                                               # Z порожня → евристика
                from_wh = resolve_wh(raw_from, service)
                if from_wh and to_wh:
                    mtype = "transfer"
                elif to_wh:
                    mtype = "receipt"
                elif from_wh:
                    mtype = "writeoff"
                else:
                    counts["skipped"] += 1
                    errors.append(f"рядок {i} «{name}»: немає складів (from/to)")
                    continue

            if mtype == "receipt" and not to_wh:
                counts["skipped"] += 1
                errors.append(f"рядок {i} «{name}»: надходження без складу-отримувача")
                continue
            if mtype == "transfer" and not (from_wh and to_wh):
                counts["skipped"] += 1
                errors.append(f"рядок {i} «{name}»: переміщення без складу (from/to)")
                continue

            instance = None
            # Серійність визначає КАРТКА (номенклатура), а не наявність серійника
            # в рядку руху: рядок може мати лише № картки (Поле 12) без колонки P.
            if nom.is_serialized:
                # матч екземпляра: спершу по номеру картки (Items№ ↔ Поле 12),
                # тоді по серійному (якщо він у рядку є).
                by_card = (db.query(Instance)
                           .filter(Instance.card_number == card,
                                   Instance.nomenclature_id == nom.id).first()) if card else None
                # Крос-перевірка лише коли серійник у рядку заданий.
                if by_card and serial and by_card.serial_no != serial:
                    counts["skipped"] += 1
                    errors.append(
                        f"рядок {i} «{name}»: картка {card} → серійний {by_card.serial_no} "
                        f"у каталозі, а в Переміщеннях {serial} — не збігається")
                    continue
                instance = by_card
                if not instance and serial:
                    instance = db.query(Instance).filter(Instance.serial_no == serial).first()
                if not instance:
                    if serial:
                        instance = Instance(nomenclature_id=nom.id, serial_no=serial,
                                            card_number=card, is_official=nom.is_official)
                        db.add(instance); db.flush()
                        counts["instances_created"] += 1
                    else:
                        counts["skipped"] += 1
                        errors.append(
                            f"рядок {i} «{name}»: серійна картка, але екземпляр за "
                            f"карткою «{card}» не знайдено (у рядку немає серійного)")
                        continue
                elif card and not instance.card_number:
                    instance.card_number = card
                qty = Decimal(1)
            else:
                qty = qin if qin > 0 else qout
                if qty <= 0:
                    counts["skipped"] += 1
                    errors.append(f"рядок {i} «{name}»: нульова кількість")
                    continue

            iso = _parse_date(mc(cells, "entry_date")) or _parse_date(mc(cells, "doc_date"))
            try:
                mdate = date_cls.fromisoformat(iso) if iso else date_cls.today()
            except ValueError:
                mdate = date_cls.today()

            doc_number = _clean(mc(cells, "doc_number"))
            mv = CustodyMovement(
                date=mdate, type=mtype, nomenclature_id=nom.id,
                from_warehouse_id=from_wh.id if from_wh else None,
                to_warehouse_id=to_wh.id if to_wh else None,
                instance_id=instance.id if instance else None,
                quantity=qty, is_official=nom.is_official,
                card_number=card, doc_number=doc_number,
                created_by=user.id,
            )
            db.add(mv)
            counts["movements"] += 1
            counts["classified_by_column" if op_hint else "by_heuristic"] += 1
            if instance:
                serial_movements.setdefault(instance, []).append(mv)
            # МВО складів із J/K (особи мають існувати; не знайдено → у звіт)
            for wh, raw_mvo in ((from_wh, mc(cells, "mvo_from")), (to_wh, mc(cells, "mvo_to"))):
                nm = _clean(raw_mvo)
                if not wh or not nm:
                    continue
                pid = _resolve_person(nm, person_lookup)
                if pid:
                    mvo_obs.setdefault(wh.id, []).append((mdate, pid))
                else:
                    mvo_unmatched.add(nm)
            if doc_number:
                key = (doc_number, mtype,
                       from_wh.id if from_wh else None,
                       to_wh.id if to_wh else None)
                grp = doc_groups.setdefault(key, {
                    "date": mdate, "movements": [],
                    "form": form, "counterparty": counterparty})
                grp["movements"].append(mv)
        except Exception as e:
            counts["skipped"] += 1
            errors.append(f"рядок {i} «{name}»: {e}")

    db.flush()   # присвоїти id рухам перед лінкуванням у документи

    # Детерміноване розміщення серійних: для кожного екземпляра — призначення
    # хронологічно ПІЗНІШОГО руху (дата → номер накладної → id), незалежно від
    # порядку рядків у файлі. Раніше «останній рядок перемагав» → при рухах в
    # одну дату екземпляр міг опинитися не на тому складі.
    for instance, mvs in serial_movements.items():
        latest = max(mvs, key=lambda m: (m.date, doc_sort_key(m.doc_number), m.id))
        instance.current_warehouse_id = latest.to_warehouse_id

    _build_mvo_journal(db, mvo_obs, counts)
    for nm in sorted(mvo_unmatched):
        errors.append(f"МВО «{nm}» не знайдено — заведіть особу до імпорту")

    _backfill_documents(db, user, doc_groups, counts)
    db.commit()
    return {**counts, "errors": errors[:100]}


def _backfill_documents(db: Session, user: User, doc_groups: dict, counts: dict) -> None:
    """Створити custody_documents з груп імпортованих рухів (по номеру накладної),
    щоб імпортовані накладні одразу були повноцінними документами (не «без документа»).
    Операція (receipt/transfer) і форма (накладна/акт) — з колонок Z/Y; контрагент
    для надходжень — із «Звідки»; статус signed (історичні)."""
    for (doc_number, mtype, from_id, to_id), g in doc_groups.items():
        operation = "receipt" if mtype == "receipt" else "transfer"
        doc = CustodyDocument(
            operation=operation, form=g.get("form") or "накладна", doc_number=doc_number,
            doc_date=g["date"].isoformat(), date_operation=g["date"].isoformat(),
            from_warehouse_id=from_id, to_warehouse_id=to_id,
            counterparty=g.get("counterparty"),
            status="signed", signed_at=datetime.utcnow(), signed_by=user.id,
            extra_data={}, created_by=user.id,
        )
        db.add(doc); db.flush()
        for mv in g["movements"]:
            mv.document_id = doc.id
        db.flush()
        snap_nakladna(doc, db)
        counts["documents_created"] += 1


# ── Видачі (assignments) — 3-й крок з файлу Items («Де» → «[Прізвище]») ───────

@router.post("/import/assignments", status_code=status.HTTP_200_OK)
def import_assignments_v2(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    """Видачі особам з колонки «Де» файлу Items («<підрозділ> [Прізвище]»).
    Запускати ПІСЛЯ Items + Переміщень (майно має бути вже розміщене). К-сть = 1;
    дата = дата накладної руху, що розмістив майно. Матч особи по прізвищу."""
    try:
        wb = load_workbook(BytesIO(file.file.read()), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не вдалось прочитати XLSX: {e}")
    ws = wb.active
    header_row = _find_header_row(ws)
    if not header_row:
        raise HTTPException(400, "Не знайдено заголовок «Назва»/«Товар»")
    cmap = _col_map(ws, header_row)

    services = {s.name.strip().lower(): s for s in db.query(Service).all()}
    noms = {(n.name.strip().lower(), n.service_id): n for n in db.query(Nomenclature).all()}
    persons = db.query(Person).all()
    # Внутрішні підрозділи, довші назви першими — для найдовшого префіксного матчу.
    internal_units = sorted(
        [u for u in db.query(Unit).all() if not u.is_external],
        key=lambda u: len(u.name or ""), reverse=True)
    wh_of_unit = {
        w.unit_id: w for w in db.query(Warehouse).filter(Warehouse.type == "unit").all()
    }

    counts = {"rows": 0, "assignments": 0, "persons_unit_set": 0, "skipped": 0}
    errors = []

    def col(row, key):
        for c, k in cmap.items():
            if k == key:
                return row[c - 1].value
        return None

    def parse_de(loc):
        """«<підрозділ> <Прізвище>» → (unit, surname). surname=None якщо лише
        підрозділ; (None, None) якщо склад/порожньо/не розпізнано."""
        s = _clean(loc)
        if not s or is_service_warehouse_location(s):
            return (None, None)
        low = s.lower()
        for u in internal_units:
            n = (u.name or "").strip()
            if not n:
                continue
            if low == n.lower():
                return (u, None)
            if low.startswith(n.lower() + " "):
                return (u, s[len(n):].strip() or None)
        return (None, None)

    def latest_move_date(inst_id=None, nom_id=None, wh_id=None):
        q = db.query(CustodyMovement)
        if inst_id:
            q = q.filter(CustodyMovement.instance_id == inst_id)
        else:
            q = q.filter(CustodyMovement.nomenclature_id == nom_id,
                         CustodyMovement.to_warehouse_id == wh_id,
                         CustodyMovement.instance_id.is_(None))
        m = q.order_by(CustodyMovement.date.desc(), CustodyMovement.id.desc()).first()
        return m.date if m else date_cls.today()

    for row in ws.iter_rows(min_row=header_row + 1):
        name = _clean(col(row, "name"))
        if not name:
            continue
        loc = col(row, "location")
        unit, surname = parse_de(loc)
        if unit is None:
            # склад/порожньо — пропускаємо тихо; текст із нерозпізнаним підрозділом — у звіт
            s = _clean(loc)
            if s and not is_service_warehouse_location(s):
                counts["rows"] += 1
                counts["skipped"] += 1
                errors.append(f"«{name}»: не розпізнано підрозділ у «Де»: «{s}»")
            continue
        if surname is None:
            continue  # лише підрозділ — не видано особі
        counts["rows"] += 1
        try:
            svc_name = _clean(col(row, "service"))
            service = services.get(svc_name.strip().lower()) if svc_name else None
            if not service:
                counts["skipped"] += 1
                errors.append(f"«{name}»: службу «{svc_name}» не знайдено")
                continue
            nom = noms.get((name.strip().lower(), service.id))
            if not nom:
                counts["skipped"] += 1
                errors.append(f"«{name}»: номенклатуру не знайдено")
                continue
            wh = wh_of_unit.get(unit.id)
            if not wh:
                counts["skipped"] += 1
                errors.append(f"«{name}»: у підрозділу «{unit.name}» немає складу")
                continue
            # Шукаємо серед ІСНУЮЧИХ осіб (люди імпортуються окремо, не створюємо
            # тут): спершу вже в цьому підрозділі, інакше — серед тих, у кого
            # підрозділ ще не заданий (їм проставимо цей підрозділ).
            sl = surname.lower()
            in_unit = [p for p in persons if (p.last_name or "").strip().lower() == sl and p.unit_id == unit.id]
            free = [p for p in persons if (p.last_name or "").strip().lower() == sl and p.unit_id is None]
            candidates = in_unit or free
            if len(candidates) > 1:
                counts["skipped"] += 1
                errors.append(f"«{name}»: неоднозначне прізвище «{surname}» "
                              f"({'у ' + unit.name if in_unit else 'без підрозділу'})")
                continue
            if not candidates:
                counts["skipped"] += 1
                errors.append(f"«{name}»: особу «{surname}» не знайдено — імпортуйте людей")
                continue
            person = candidates[0]
            if person.unit_id is None:      # оновлюємо підрозділ, якщо не заданий
                person.unit_id = unit.id
                counts["persons_unit_set"] += 1

            if nom.is_serialized:
                card = _clean(col(row, "card_number"))
                inst = (db.query(Instance).filter(Instance.card_number == card,
                        Instance.nomenclature_id == nom.id).first()) if card else None
                if not inst:
                    counts["skipped"] += 1
                    errors.append(f"«{name}»: екземпляр за карткою «{card}» не знайдено")
                    continue
                if inst.current_warehouse_id != wh.id:
                    counts["skipped"] += 1
                    errors.append(f"«{name}» ({inst.serial_no}): не на складі підрозділу «{unit.name}»")
                    continue
                if db.query(Assignment).filter(Assignment.instance_id == inst.id,
                                               Assignment.returned_date.is_(None)).first():
                    counts["skipped"] += 1
                    errors.append(f"«{name}» ({inst.serial_no}): вже видано")
                    continue
                db.add(Assignment(
                    warehouse_id=wh.id, person_id=person.id, nomenclature_id=nom.id,
                    instance_id=inst.id, quantity=Decimal(1), is_official=inst.is_official,
                    issued_date=latest_move_date(inst_id=inst.id), created_by=user.id))
            else:
                free = balance_of(db, wh.id, nom.id, nom.is_official) - _active_issued(db, wh.id, nom.id, nom.is_official)
                if free < 1:
                    counts["skipped"] += 1
                    errors.append(f"«{name}»: недостатньо на складі «{unit.name}» (вільно {free})")
                    continue
                db.add(Assignment(
                    warehouse_id=wh.id, person_id=person.id, nomenclature_id=nom.id,
                    instance_id=None, quantity=Decimal(1), is_official=nom.is_official,
                    issued_date=latest_move_date(nom_id=nom.id, wh_id=wh.id), created_by=user.id))
            counts["assignments"] += 1
        except Exception as e:
            counts["skipped"] += 1
            errors.append(f"«{name}»: {e}")

    db.commit()
    return {**counts, "errors": errors[:100]}
