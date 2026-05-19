import io
import os
from datetime import datetime, timedelta
from decimal import Decimal
from typing import List, Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.database import get_db
from app.invoice_export import (
    ROW_TOTALS, ROW_CHIEF_POST, ROW_CHIEF_NAME, ROW_QTY_WORDS, ROW_AMT_WORDS,
    ROW_SENDER, ROW_RECEIVER, ROW_FIN_POST, ROW_FIN_NAME,
    TEMPLATE_LAST_ROW, adjust_item_rows,
)
from app.models import (
    Document, DocumentItem, Item, Movement, OpType, Person, Service,
    UnitSettings, User,
)
from app.uk_num2words import amount_to_words_uk, qty_to_words_uk

router = APIRouter(prefix="/api/documents", tags=["documents"])

DOC_TYPES = ("надходження", "переміщення", "накладна_25")

REQUIRED_FIELDS = {
    "надходження":  ["doc_number", "doc_date", "to_unit"],
    "переміщення":  ["doc_number", "doc_date", "from_unit", "to_unit"],
    "накладна_25":  ["doc_number", "doc_date"],
}

# Keys stored inside extra_data JSON
SNAP_KEYS = [
    "snap_unit_name", "snap_edrpou", "composed_location",
    "snap_op_type_name", "snap_service_name",
    "snap_service_chief_post", "snap_service_chief_name",
    "snap_sender_subdiv", "snap_sender_post", "snap_sender_name",
    "snap_recv_subdiv", "snap_recv_rank", "snap_recv_name", "snap_recv_post",
    "snap_fin_post", "snap_fin_name",
    "validity_date", "total_qty_words", "total_amount_words",
]

UK_MONTHS = ["січня", "лютого", "березня", "квітня", "травня", "червня",
             "липня", "серпня", "вересня", "жовтня", "листопада", "грудня"]


# ── Schemas ────────────────────────────────────────────────────────────────

class DocItemIn(BaseModel):
    sort_order: Optional[int] = None
    item_id: Optional[int] = None              # FK → items.id (накладна_25 snap source)
    item_name: Optional[str] = None
    nomenclature_code: Optional[str] = None
    unit_of_measure: Optional[str] = None
    category: Optional[str] = None
    quantity: Optional[Decimal] = None         # qty_sent
    qty_received: Optional[Decimal] = None
    price: Optional[Decimal] = None
    amount: Optional[Decimal] = None
    notes: Optional[str] = None


class DocIn(BaseModel):
    doc_type: str = "накладна_25"
    doc_number: Optional[str] = None
    doc_date: Optional[str] = None
    date_operation: Optional[str] = None
    from_unit: Optional[str] = None
    to_unit: Optional[str] = None
    basis: Optional[str] = None
    op_type_id: Optional[int] = None
    service_id: Optional[int] = None
    sender_id: Optional[int] = None
    receiver_id: Optional[int] = None
    fin_id: Optional[int] = None
    items: List[DocItemIn] = []


# ── Helpers ────────────────────────────────────────────────────────────────

def _person_full_name(p: Person) -> str:
    """«Ім'я ПРІЗВИЩЕ» — first_name + last_name.upper()"""
    return " ".join(filter(None, [
        (p.first_name or "").strip(),
        (p.last_name or "").strip().upper(),
    ]))


def _parse_date(s: Optional[str]):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _calc_validity(doc_date: Optional[str]) -> str:
    """`doc_date + 3 days` formatted as `"DD" місяця YYYY року`."""
    dt = _parse_date(doc_date)
    if not dt:
        return ""
    valid = dt + timedelta(days=3)
    return f'"{valid.day:02d}" {UK_MONTHS[valid.month - 1]} {valid.year} року'


def _next_doc_number(prefix: str, db: Session, exclude_id: Optional[int] = None) -> str:
    """`prefix` + (max trailing integer among existing docs with that prefix + 1)."""
    if not prefix:
        return ""
    q = db.query(Document.doc_number).filter(
        Document.doc_number.like(prefix + "%"),
    )
    if exclude_id is not None:
        q = q.filter(Document.id != exclude_id)
    max_n = 0
    for (num,) in q.all():
        if not num or not num.startswith(prefix):
            continue
        try:
            n = int(num[len(prefix):])
            if n > max_n:
                max_n = n
        except ValueError:
            pass
    return f"{prefix}{max_n + 1}"


def _snap_nakladna(doc: Document, db: Session):
    """Refresh snapshot fields on `doc` from current FK references.

    Called on every save while status == 'draft'. Stored values are read at
    export time (signed docs are immutable so their snap is frozen).
    """
    extra = dict(doc.extra_data or {})
    # Clear any prior snap, keep other keys untouched
    for k in SNAP_KEYS:
        extra.pop(k, None)

    settings = db.query(UnitSettings).first()
    if settings:
        extra["snap_unit_name"]    = settings.name or ""
        extra["snap_edrpou"]       = settings.edrpou or ""
        extra["composed_location"] = settings.location or ""

    if doc.op_type_id:
        ot = db.get(OpType, doc.op_type_id)
        if ot:
            extra["snap_op_type_name"] = ot.name or ""

    if doc.service_id:
        sv = db.get(Service, doc.service_id)
        if sv:
            doc.service = sv.name or ""
            extra["snap_service_name"]       = sv.name or ""
            extra["snap_service_chief_post"] = sv.chief_position or ""
            extra["snap_service_chief_name"] = sv.chief_name or ""

    if doc.sender_id:
        p = db.get(Person, doc.sender_id)
        if p:
            doc.from_unit = p.unit or ""
            extra["snap_sender_subdiv"] = p.unit or ""
            extra["snap_sender_post"]   = p.position or ""
            extra["snap_sender_name"]   = _person_full_name(p)

    if doc.receiver_id:
        p = db.get(Person, doc.receiver_id)
        if p:
            doc.to_unit = p.unit or ""
            extra["snap_recv_subdiv"] = p.unit or ""
            extra["snap_recv_rank"]   = p.rank or ""
            extra["snap_recv_name"]   = _person_full_name(p)
            extra["snap_recv_post"]   = p.position or ""

    if doc.fin_id:
        p = db.get(Person, doc.fin_id)
        if p:
            extra["snap_fin_post"] = p.position or ""
            extra["snap_fin_name"] = _person_full_name(p)

    extra["validity_date"] = _calc_validity(doc.doc_date)

    total_qty = Decimal(0)
    total_amt = Decimal(0)
    for it in doc.items:
        q = Decimal(str(it.quantity or 0))
        p = Decimal(str(it.price or 0))
        total_qty += q
        total_amt += q * p
    extra["total_qty_words"]    = qty_to_words_uk(total_qty) if total_qty else ""
    extra["total_amount_words"] = amount_to_words_uk(total_amt) if total_amt else ""

    doc.extra_data = extra


def _apply_items(doc: Document, items: List[DocItemIn], db: Session, snap_from_items: bool):
    """Replace doc.items. For накладна_25 (`snap_from_items=True`), missing snap
    fields are filled from the items table using `item_id`.
    """
    for old in list(doc.items):
        db.delete(old)
    db.flush()
    for idx, src in enumerate(items):
        snap_name = src.item_name
        snap_code = src.nomenclature_code
        snap_unit = src.unit_of_measure
        snap_cat  = src.category
        snap_price = src.price

        if snap_from_items and src.item_id:
            ref = db.get(Item, src.item_id)
            if ref:
                snap_name  = ref.name
                snap_code  = ref.nomenclature_code
                snap_unit  = ref.unit_of_measure
                snap_cat   = ref.category
                snap_price = ref.price

        qty = src.quantity
        amount = src.amount
        if amount is None and qty is not None and snap_price is not None:
            amount = Decimal(str(qty)) * Decimal(str(snap_price))

        db.add(DocumentItem(
            document=doc,
            sort_order=src.sort_order if src.sort_order is not None else idx,
            item_id=src.item_id,
            item_name=snap_name,
            nomenclature_code=snap_code,
            unit_of_measure=snap_unit,
            category=snap_cat,
            quantity=qty,
            qty_received=src.qty_received,
            price=snap_price,
            amount=amount,
            notes=src.notes,
        ))
    db.flush()


def _sorted_items(doc: Document):
    return sorted(doc.items, key=lambda x: (x.sort_order or 0, x.id))


def _items_for_display(doc: Document) -> list:
    if doc.items:
        return [
            {
                "id": it.id,
                "sort_order": it.sort_order,
                "item_id": it.item_id,
                "item_name": it.item_name,
                "nomenclature_code": it.nomenclature_code,
                "unit_of_measure": it.unit_of_measure,
                "category": it.category,
                "quantity": float(it.quantity) if it.quantity is not None else None,
                "qty_received": float(it.qty_received) if it.qty_received is not None else None,
                "price": float(it.price) if it.price is not None else None,
                "amount": float(it.amount) if it.amount is not None else None,
                "notes": it.notes,
            }
            for it in _sorted_items(doc)
        ]
    # Fallback for imported docs (built from movements)
    return [
        {
            "id": None,
            "sort_order": i,
            "item_name": m.item_name,
            "nomenclature_code": m.nomenclature_code,
            "unit_of_measure": m.unit_of_measure,
            "category": str(m.category) if m.category is not None else None,
            "quantity": float(m.qty_in or m.qty_out or 0),
            "qty_received": None,
            "price": float(m.price) if m.price is not None else None,
            "amount": float(m.price * (m.qty_in or m.qty_out or 0)) if m.price else None,
            "notes": m.serial_number,
        }
        for i, m in enumerate(sorted(doc.movements, key=lambda x: x.id), 1)
    ]


def _doc_to_dict(doc: Document) -> dict:
    extra = doc.extra_data or {}
    return {
        "id": doc.id,
        "doc_type": doc.doc_type,
        "doc_type_label": doc.movements[0].doc_type if doc.movements else None,
        "doc_number": doc.doc_number,
        "doc_date": doc.doc_date,
        "date_operation": doc.date_operation,
        "from_unit": doc.from_unit,
        "to_unit": doc.to_unit,
        "basis": doc.basis,
        "service": doc.service,
        "op_type_id": doc.op_type_id,
        "service_id": doc.service_id,
        "sender_id": doc.sender_id,
        "receiver_id": doc.receiver_id,
        "fin_id": doc.fin_id,
        "status": doc.status,
        "signed_at": doc.signed_at.isoformat() if doc.signed_at else None,
        "extra_data": extra,
        "items": _items_for_display(doc),
    }


def _get_or_404(doc_id: int, db: Session) -> Document:
    doc = db.query(Document).filter(Document.id == doc_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Not found")
    return doc


def _resolve_auto_number(doc: Document, db: Session):
    """If `doc.doc_number` is blank and op_type has a `number_prefix`, fill it
    via next sequential number. Warns on duplicate but does not block.
    Returns a list of warnings.
    """
    warnings = []
    if doc.doc_number:
        # Duplicate check
        dup = db.query(Document).filter(
            Document.doc_number == doc.doc_number,
            Document.id != (doc.id or -1),
        ).first()
        if dup:
            warnings.append(f"Накладна з номером {doc.doc_number} вже існує")
        return warnings

    if doc.op_type_id:
        ot = db.get(OpType, doc.op_type_id)
        if ot and ot.number_prefix:
            doc.doc_number = _next_doc_number(ot.number_prefix, db, exclude_id=doc.id)
    return warnings


# ── CRUD ───────────────────────────────────────────────────────────────────

@router.get("")
def list_documents(
    doc_type: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(Document)
    if doc_type:
        q = q.filter(Document.doc_type == doc_type)
    docs = q.order_by(Document.created_at.desc()).all()
    return [
        {
            "id": d.id,
            "doc_type": d.doc_type,
            "doc_type_label": d.movements[0].doc_type if d.movements else None,
            "doc_number": d.doc_number,
            "doc_date": d.doc_date,
            "from_unit": d.from_unit,
            "to_unit": d.to_unit,
            "status": d.status,
            "items_count": len(d.movements) if d.status != "draft" else len(d.items),
        }
        for d in docs
    ]


@router.get("/{doc_id}")
def get_document(doc_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    return _doc_to_dict(_get_or_404(doc_id, db))


@router.post("", status_code=status.HTTP_201_CREATED)
def create_document(payload: DocIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if payload.doc_type not in DOC_TYPES:
        raise HTTPException(400, f"doc_type must be one of: {DOC_TYPES}")

    doc = Document(
        doc_type=payload.doc_type,
        doc_number=payload.doc_number,
        doc_date=payload.doc_date,
        date_operation=payload.date_operation or payload.doc_date,
        from_unit=payload.from_unit,
        to_unit=payload.to_unit,
        basis=payload.basis,
        op_type_id=payload.op_type_id,
        service_id=payload.service_id,
        sender_id=payload.sender_id,
        receiver_id=payload.receiver_id,
        fin_id=payload.fin_id,
        status="draft",
        extra_data={},
        created_by=user.id,
    )
    db.add(doc)
    db.flush()

    warnings = _resolve_auto_number(doc, db)

    is_nakl = payload.doc_type == "накладна_25"
    _apply_items(doc, payload.items, db, snap_from_items=is_nakl)

    if is_nakl:
        _snap_nakladna(doc, db)

    db.commit()
    db.refresh(doc)
    out = _doc_to_dict(doc)
    if warnings:
        out["warnings"] = warnings
    return out


@router.put("/{doc_id}")
def update_document(doc_id: int, payload: DocIn, db: Session = Depends(get_db), _=Depends(get_current_user)):
    doc = _get_or_404(doc_id, db)
    if doc.status != "draft":
        raise HTTPException(400, "Підписаний документ не можна редагувати. Спочатку зніміть підпис.")

    doc.doc_type = payload.doc_type
    doc.doc_number = payload.doc_number
    doc.doc_date = payload.doc_date
    doc.date_operation = payload.date_operation or payload.doc_date
    doc.from_unit = payload.from_unit
    doc.to_unit = payload.to_unit
    doc.basis = payload.basis
    doc.op_type_id = payload.op_type_id
    doc.service_id = payload.service_id
    doc.sender_id = payload.sender_id
    doc.receiver_id = payload.receiver_id
    doc.fin_id = payload.fin_id

    warnings = _resolve_auto_number(doc, db)

    is_nakl = payload.doc_type == "накладна_25"
    _apply_items(doc, payload.items, db, snap_from_items=is_nakl)

    if is_nakl:
        _snap_nakladna(doc, db)

    db.commit()
    db.refresh(doc)
    out = _doc_to_dict(doc)
    if warnings:
        out["warnings"] = warnings
    return out


@router.delete("/{doc_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_document(doc_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    doc = _get_or_404(doc_id, db)
    if doc.status != "draft":
        raise HTTPException(400, "Не можна видалити підписаний документ. Спочатку зніміть підпис.")
    db.delete(doc)
    db.commit()


# ── Sign / Unsign ──────────────────────────────────────────────────────────

@router.post("/{doc_id}/sign")
def sign_document(doc_id: int, db: Session = Depends(get_db), user=Depends(get_current_user)):
    doc = _get_or_404(doc_id, db)
    if doc.status != "draft":
        raise HTTPException(400, "Документ вже підписано.")

    errors = []
    for field in REQUIRED_FIELDS.get(doc.doc_type, []):
        if not getattr(doc, field, None):
            errors.append(field)
    if not doc.items:
        errors.append("items (список позицій порожній)")
    if errors:
        raise HTTPException(422, {"detail": "Заповніть обов'язкові поля", "missing": errors})

    for it in _sorted_items(doc):
        is_incoming = (doc.doc_type == "надходження")
        m = Movement(
            document_id=doc.id,
            entry_date=doc.doc_date,
            item_name=it.item_name,
            unit_of_measure=it.unit_of_measure,
            category=it.category,
            nomenclature_code=it.nomenclature_code,
            qty_in=it.quantity if is_incoming else (it.qty_received or None),
            qty_out=None if is_incoming else it.quantity,
            from_unit=doc.from_unit,
            to_unit=doc.to_unit,
            basis=doc.basis,
            doc_date=doc.doc_date,
            doc_number=doc.doc_number,
            doc_type=doc.doc_type,
            service=doc.service,
            price=it.price,
            mvo_from_id=doc.sender_id,
            mvo_to_id=doc.receiver_id,
            created_by=user.id,
        )
        db.add(m)

    doc.status = "signed"
    doc.signed_at = datetime.utcnow()
    doc.signed_by = user.id
    db.commit()
    db.refresh(doc)
    return _doc_to_dict(doc)


@router.post("/{doc_id}/unsign")
def unsign_document(doc_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    doc = _get_or_404(doc_id, db)
    if doc.status == "draft":
        raise HTTPException(400, "Документ вже у статусі чернетки.")

    db.query(Movement).filter(Movement.document_id == doc.id).delete(synchronize_session=False)
    doc.status = "draft"
    doc.signed_at = None
    doc.signed_by = None
    db.commit()
    db.refresh(doc)
    return _doc_to_dict(doc)


# ── Excel export (накладна_25) ─────────────────────────────────────────────

EXPORT_REQUIRED_SNAP = ("snap_unit_name",)  # if missing → 400, doc not migrated


def _has_snap(doc: Document) -> bool:
    extra = doc.extra_data or {}
    return any(extra.get(k) for k in EXPORT_REQUIRED_SNAP)


@router.get("/{doc_id}/export/xlsx")
def export_xlsx(doc_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    doc = _get_or_404(doc_id, db)
    if doc.doc_type != "накладна_25":
        raise HTTPException(400, "XLSX-експорт доступний лише для типу «накладна_25»")
    if not _has_snap(doc):
        raise HTTPException(
            400,
            "Документ створено до переходу на снапшот-архітектуру. "
            "Відкрийте його в редакторі та збережіть, щоб згенерувати snap-поля.",
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    extra = doc.extra_data or {}
    items_list = _items_for_display(doc)
    n_items = len(items_list)

    # Load template
    tpl_path = os.path.join(os.path.dirname(__file__), "..", "nakladna_template.xlsx")
    wb = load_workbook(tpl_path)
    ws = wb.active

    # ── Handle variable N (delegated to invoice_export module) ────────────
    shift = adjust_item_rows(ws, n_items)

    # ── Helpers for writing ───────────────────────────────────────────────
    def sv(addr, value):
        ws[addr].value = value

    def sr(row, col, value):
        ws.cell(row=row, column=col).value = value

    # ── Header (rows 1–15, fixed) ─────────────────────────────────────────
    sv("B2", extra.get("snap_unit_name") or "")
    sv("D4", extra.get("snap_edrpou") or "")
    sv("B6", extra.get("validity_date") or 'Дійсна до "____" _________ ____ року')
    sv("E7", doc.doc_number or "")
    sv("I8",  extra.get("composed_location") or "")
    sv("I10", doc.doc_date or "")
    sr(12, 3, doc.date_operation or doc.doc_date or "")
    sr(12, 9, extra.get("snap_service_name") or doc.service or "")
    sv("C13", extra.get("snap_op_type_name") or "")
    sv("I13", doc.basis or "")

    recv_full = " ".join(filter(None, [
        (extra.get("snap_recv_rank") or "").strip(),
        (extra.get("snap_recv_name") or "").strip(),
    ]))
    sv("C14", recv_full)

    sv("C15", extra.get("snap_sender_subdiv") or doc.from_unit or "")
    sv("I15", extra.get("snap_recv_subdiv") or doc.to_unit or "")

    # ── Item rows ─────────────────────────────────────────────────────────
    total_qty = Decimal(0)
    total_amt = Decimal(0)
    for idx, it in enumerate(items_list):
        r = 20 + idx
        qty = Decimal(str(it["quantity"])) if it["quantity"] is not None else Decimal(0)
        price = Decimal(str(it["price"])) if it["price"] is not None else Decimal(0)
        amt = qty * price
        sr(r, 1,  idx + 1)
        sr(r, 2,  it["item_name"] or "")
        sr(r, 3,  it["nomenclature_code"] or "")
        sr(r, 4,  it["unit_of_measure"] or "")
        sr(r, 5,  it["category"] or "")
        sr(r, 6,  float(price) if price else "")
        sr(r, 7,  float(qty) if qty else "")
        sr(r, 8,  float(it["qty_received"]) if it["qty_received"] is not None else "")
        sr(r, 9,  float(amt) if amt else "")
        sr(r, 10, it["notes"] or "")
        total_qty += qty
        total_amt += amt

    # ── Totals (row 25 + shift) ───────────────────────────────────────────
    sr(ROW_TOTALS + shift, 7, float(total_qty) if total_qty else "")
    sr(ROW_TOTALS + shift, 8, float(total_qty) if total_qty else "")
    sr(ROW_TOTALS + shift, 9, float(total_amt) if total_amt else "")

    # ── Service chief signature (row 27 + shift) ──────────────────────────
    sr(ROW_CHIEF_POST + shift, 1,  extra.get("snap_service_chief_post") or "")
    sr(ROW_CHIEF_NAME + shift, 10, extra.get("snap_service_chief_name") or "")

    # ── Total in words ────────────────────────────────────────────────────
    sr(ROW_QTY_WORDS + shift, 3, extra.get("total_qty_words") or "")  # C29:I29
    amt_words = extra.get("total_amount_words") or ""
    sr(ROW_AMT_WORDS + shift, 1, f"на\xa0суму {amt_words}" if amt_words else "")  # A31:J31

    # ── МВО ───────────────────────────────────────────────────────────────
    sr(ROW_SENDER + shift, 1,  extra.get("snap_sender_post") or "")
    sr(ROW_SENDER + shift, 10, extra.get("snap_sender_name") or "")
    sr(ROW_RECEIVER + shift, 1,  extra.get("snap_recv_post") or "")
    sr(ROW_RECEIVER + shift, 10, extra.get("snap_recv_name") or "")

    # ── Fin signature (rows 41–43 + shift) ────────────────────────────────
    sr(ROW_FIN_POST + shift, 1,  extra.get("snap_fin_post") or "")
    sr(ROW_FIN_NAME + shift, 10, extra.get("snap_fin_name") or "")

    # ── Update print area ─────────────────────────────────────────────────
    new_last = TEMPLATE_LAST_ROW + shift
    ws.print_area = f"A1:J{new_last}"

    # ── Stream ────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    display_name = f"накладна_{doc.doc_number or doc.id}.xlsx"
    encoded = quote(display_name.encode("utf-8"))
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition":
                f'attachment; filename="nakladna.xlsx"; filename*=UTF-8\'\'{encoded}'
        },
    )
