"""Shared Додаток-25 rendering + snapshot helpers.

Pure (no DB models): reads snap-dict fields off a duck-typed document object and
formats dates/names. Used by BOTH the v2 `custody_*` chain and (until it is
retired) the v1 `document_*` chain, so neither layer depends on the other.
"""
import io
import os
from datetime import datetime, timedelta
from decimal import Decimal
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from app.invoice_export import (
    ROW_TOTALS, ROW_CHIEF_POST, ROW_CHIEF_NAME, ROW_QTY_WORDS, ROW_AMT_WORDS,
    ROW_SENDER, ROW_RECEIVER, ROW_FIN_POST, ROW_FIN_NAME,
    TEMPLATE_LAST_ROW, adjust_item_rows,
)


# Every snap-text key that lives in a document's `extra_data` JSON.
SNAP_KEYS = [
    "snap_unit_name", "snap_edrpou", "composed_location",
    "snap_op_type_name", "snap_service_name",
    "snap_service_chief_post", "snap_service_chief_name",
    "snap_sender_subdiv", "snap_sender_post", "snap_sender_name",
    "snap_recv_subdiv", "snap_recv_rank", "snap_recv_name", "snap_recv_post",
    "snap_fin_post", "snap_fin_name",
    "validity_date", "total_qty_words", "total_amount_words",
]

# A signed/migrated doc must have at least these keys in extra_data;
# missing snap → 400 «потрібна snap-міграція».
EXPORT_REQUIRED_SNAP: Tuple[str, ...] = ("snap_unit_name",)

UK_MONTHS = ["січня", "лютого", "березня", "квітня", "травня", "червня",
             "липня", "серпня", "вересня", "жовтня", "листопада", "грудня"]


# ── Date / name helpers ───────────────────────────────────────────────────

def person_full_name(p) -> str:
    """«Ім'я ПРІЗВИЩЕ» — first_name + last_name.upper()"""
    return " ".join(filter(None, [
        (p.first_name or "").strip(),
        (p.last_name or "").strip().upper(),
    ]))


def parse_date(s: Optional[str]):
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def calc_validity(doc_date: Optional[str]) -> str:
    """`doc_date + 3 days` formatted as `"DD" місяця YYYY року`."""
    dt = parse_date(doc_date)
    if not dt:
        return ""
    valid = dt + timedelta(days=3)
    return f'"{valid.day:02d}" {UK_MONTHS[valid.month - 1]} {valid.year} року'


# ── XLSX renderer ─────────────────────────────────────────────────────────

def has_snap(doc) -> bool:
    """A document is exportable once its snap has at least the required keys."""
    extra = doc.extra_data or {}
    return any(extra.get(k) for k in EXPORT_REQUIRED_SNAP)


def build_xlsx(doc, items_list: List[dict]) -> bytes:
    """Render `doc` (duck-typed: reads snap `extra_data` + header attrs) + its
    `items_list` (in display order) into Додаток-25 XLSX bytes."""
    extra = doc.extra_data or {}

    tpl_path = os.path.join(os.path.dirname(__file__), "nakladna_template.xlsx")
    wb = load_workbook(tpl_path)
    ws = wb.active

    shift = adjust_item_rows(ws, len(items_list))

    def sv(addr, value):
        ws[addr].value = value

    def sr(row, col, value):
        ws.cell(row=row, column=col).value = value

    # ── Header (rows 1–15, fixed) ─────────────────────────────────────────
    sv("B2", extra.get("snap_unit_name") or "")
    sv("D4", extra.get("snap_edrpou") or "")
    sv("B6", extra.get("validity_date") or 'Дійсна до "____" _________ ____ року')
    sv("E7", doc.doc_number or "")
    sv("I8", extra.get("composed_location") or "")
    sv("I10", doc.doc_date or "")
    sr(12, 3, doc.date_operation or doc.doc_date or "")
    sr(12, 9, extra.get("snap_service_name") or doc.service or "")
    sv("C13", extra.get("snap_op_type_name") or "")
    sv("I13", doc.basis or "")

    recv_full = " ".join(filter(None, [
        (extra.get("snap_recv_rank") or "").strip(),
        (extra.get("snap_recv_name") or "").strip(),
    ]))
    sv("C14", recv_full)

    sv("C15", extra.get("snap_sender_subdiv") or doc.from_unit or "")
    sv("I15", extra.get("snap_recv_subdiv") or doc.to_unit or "")

    # ── Item rows ─────────────────────────────────────────────────────────
    total_qty = Decimal(0)
    total_amt = Decimal(0)
    for idx, it in enumerate(items_list):
        r = 20 + idx
        qty   = Decimal(str(it["quantity"])) if it["quantity"] is not None else Decimal(0)
        price = Decimal(str(it["price"]))    if it["price"]    is not None else Decimal(0)
        amt   = qty * price
        sr(r, 1,  idx + 1)
        sr(r, 2,  it["item_name"] or "")
        sr(r, 3,  it["nomenclature_code"] or "")
        sr(r, 4,  it["unit_of_measure"] or "")
        sr(r, 5,  it["category"] or "")
        sr(r, 6,  float(price) if price else "")
        sr(r, 7,  float(qty) if qty else "")
        sr(r, 8,  float(it["qty_received"]) if it["qty_received"] is not None else "")
        sr(r, 9,  float(amt) if amt else "")
        sr(r, 10, it["notes"] or "")
        total_qty += qty
        total_amt += amt

    # ── Totals (row 25 + shift) ───────────────────────────────────────────
    sr(ROW_TOTALS + shift, 7, float(total_qty) if total_qty else "")
    sr(ROW_TOTALS + shift, 8, float(total_qty) if total_qty else "")
    sr(ROW_TOTALS + shift, 9, float(total_amt) if total_amt else "")

    # ── Service chief signature (row 27 + shift) ──────────────────────────
    sr(ROW_CHIEF_POST + shift, 1,  extra.get("snap_service_chief_post") or "")
    sr(ROW_CHIEF_NAME + shift, 10, extra.get("snap_service_chief_name") or "")

    # ── Total in words ────────────────────────────────────────────────────
    sr(ROW_QTY_WORDS + shift, 3, extra.get("total_qty_words") or "")  # C29:I29
    amt_words = extra.get("total_amount_words") or ""
    sr(ROW_AMT_WORDS + shift, 1, f"на\xa0суму {amt_words}" if amt_words else "")  # A31:J31

    # ── МВО ───────────────────────────────────────────────────────────────
    sr(ROW_SENDER + shift, 1,    extra.get("snap_sender_post") or "")
    sr(ROW_SENDER + shift, 10,   extra.get("snap_sender_name") or "")
    sr(ROW_RECEIVER + shift, 1,  extra.get("snap_recv_post") or "")
    sr(ROW_RECEIVER + shift, 10, extra.get("snap_recv_name") or "")

    # ── Fin signature (rows 41–43 + shift) ────────────────────────────────
    sr(ROW_FIN_POST + shift, 1,  extra.get("snap_fin_post") or "")
    sr(ROW_FIN_NAME + shift, 10, extra.get("snap_fin_name") or "")

    ws.print_area = f"A1:J{TEMPLATE_LAST_ROW + shift}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
