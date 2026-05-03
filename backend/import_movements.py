#!/usr/bin/env python3
"""
Import movements from Excel file.

Usage:
    docker compose exec backend python import_movements.py /app/movements.xlsx
    docker compose exec backend python import_movements.py /app/movements.xlsx --dry-run
"""

import sys
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

sys.path.insert(0, '/app')
from app.database import SessionLocal
from app.models import Movement, OpType, Person

COLUMN_MAP = {
    'Дата внесення':           'entry_date',
    'Найменування':            'item_name',
    '№ картки':                'item_card_num',
    'Одиниця виміру':          'unit_of_measure',
    'Категорія':               'category',
    'Надійшло':                'qty_in',
    'Вибуло':                  'qty_out',
    'Звідки':                  'from_unit',
    'Куди':                    'to_unit',
    'МВО звідки':              'mvo_from_name',
    'МВО куди':                'mvo_to_name',
    'Підстава':                'basis',
    'Дата документа':          'doc_date',
    '№ документа':             'doc_number',
    'Примітка':                'serial_number',   # містить серійний номер
    'Виконавець':              'executor_name',
    'Код номенклатури':        'nomenclature_code',
    'Ціна':                    'price',
    'Служба':                  'service',
    'Поле 10':                 'op_type_detail_name',
    'номер картки':            'item_card_num_alt', # Поле 12 — fallback
    'Тип документа':           'doc_type',
    'Запасне поле2':           'op_type_name',
    'Запасне поле3':           'recipient_category',
}

SKIP_COLUMNS = {'№ без префіксу', 'Поле 11', 'Термін дії накладної'}


def parse_decimal(val) -> Decimal | None:
    if val is None:
        return None
    s = str(val).strip().replace('\xa0', '').replace(' ', '').replace(',', '.')
    if s in ('', '-', '—', 'б/н', 'б/н.'):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def parse_date(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s:
        return None
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s


def str_val(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def find_person(db, name: str | None) -> int | None:
    if not name:
        return None
    name = name.strip()
    # Шукаємо по search_name або last_name
    p = db.query(Person).filter(
        (Person.search_name.ilike(f'%{name}%')) |
        (Person.last_name.ilike(f'%{name}%'))
    ).first()
    if p:
        return p.id
    print(f"    ! Особу не знайдено: «{name}» — пропускаємо FK")
    return None


def find_op_type(db, name: str | None) -> int | None:
    if not name:
        return None
    name = name.strip()
    ot = db.query(OpType).filter(OpType.name.ilike(f'%{name}%')).first()
    if ot:
        return ot.id
    return None


def find_header_row(ws):
    for row in ws.iter_rows(max_row=30):
        for cell in row:
            if cell.value and str(cell.value).strip() == 'Найменування':
                return cell.row
    return None


def build_col_map(ws, header_row: int) -> dict:
    col_map = {}
    for cell in ws[header_row]:
        raw = str(cell.value).strip() if cell.value else ''
        if raw in COLUMN_MAP:
            col_map[cell.column] = COLUMN_MAP[raw]
    return col_map


def main():
    args = sys.argv[1:]
    if not args or args[0] in ('-h', '--help'):
        print(__doc__)
        sys.exit(0)

    filepath = args[0]
    dry_run = '--dry-run' in args

    if dry_run:
        print('=== DRY RUN — нічого не записується ===\n')

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    if not header_row:
        print("ПОМИЛКА: не знайдено рядок заголовків (шукаю колонку 'Найменування')")
        sys.exit(1)

    col_map = build_col_map(ws, header_row)
    print(f"Знайдено колонок: {len(col_map)}: {list(col_map.values())}")
    print(f"Дані починаються з рядка {header_row + 1}\n")

    db = SessionLocal()
    created = skipped = errors = 0

    try:
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            vals = {}
            for cell in row:
                if cell.column in col_map:
                    vals[col_map[cell.column]] = cell.value

            item_name = str_val(vals.get('item_name'))
            if not item_name:
                continue

            # № картки: спочатку основна колонка, якщо порожньо — Поле 12
            card_num = str_val(vals.get('item_card_num')) or str_val(vals.get('item_card_num_alt'))

            entry_date = parse_date(vals.get('entry_date'))
            doc_date   = parse_date(vals.get('doc_date'))

            if dry_run:
                print(f"  DRY  р.{row_idx:>4}  {item_name[:55]}")
                created += 1
                continue

            try:
                m = Movement(
                    entry_date        = entry_date,
                    item_name         = item_name,
                    item_card_num     = card_num,
                    unit_of_measure   = str_val(vals.get('unit_of_measure')),
                    category          = str_val(vals.get('category')),
                    qty_in            = parse_decimal(vals.get('qty_in')),
                    qty_out           = parse_decimal(vals.get('qty_out')),
                    from_unit         = str_val(vals.get('from_unit')),
                    to_unit           = str_val(vals.get('to_unit')),
                    mvo_from_id       = find_person(db, str_val(vals.get('mvo_from_name'))),
                    mvo_to_id         = find_person(db, str_val(vals.get('mvo_to_name'))),
                    executor_id       = find_person(db, str_val(vals.get('executor_name'))),
                    basis             = str_val(vals.get('basis')),
                    doc_date          = doc_date,
                    doc_number        = str_val(vals.get('doc_number')),
                    doc_type          = str_val(vals.get('doc_type')),
                    serial_number     = str_val(vals.get('serial_number')),
                    nomenclature_code = str_val(vals.get('nomenclature_code')),
                    price             = parse_decimal(vals.get('price')),
                    service           = str_val(vals.get('service')),
                    op_type_id        = find_op_type(db, str_val(vals.get('op_type_name'))),
                    recipient_category= str_val(vals.get('recipient_category')),
                    notes             = str_val(vals.get('op_type_detail_name')),
                )
                db.add(m)
                db.commit()
                print(f"  OK   р.{row_idx:>4}  {item_name[:55]}")
                created += 1
            except Exception as e:
                db.rollback()
                print(f"  ERR  р.{row_idx:>4}  {item_name[:55]}  → {e}")
                errors += 1

    finally:
        db.close()

    print(f"\n{'[DRY RUN] ' if dry_run else ''}Готово: {created} додано, {skipped} пропущено, {errors} помилок")


if __name__ == '__main__':
    main()
